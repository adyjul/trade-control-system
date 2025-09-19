# live_bot_dual_entry_liveclose.py
import asyncio
import math
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import List, Dict, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from binance import AsyncClient, BinanceSocketManager
from binance.enums import SIDE_BUY, SIDE_SELL, ORDER_TYPE_MARKET
from dotenv import load_dotenv

# ---------------- Config ----------------
@dataclass
class BotConfig:
    api_key: str = os.getenv('BINANCE_API_KEY')           # <-- set your API key here
    api_secret: str = os.getenv('BINANCE_API_SECRET')      # <-- set your API secret here
    pair: str = "AVAXUSDT"
    interval: str = "1m"
    initial_balance: float = 20.0   # assumed in USDT
    leverage: float = 20.0
    fee_rate: float = 0.0004        # taker fee default (0.04%)
    min_atr: float = 0.0005
    atr_period: int = 14
    level_mult: float = 0.2
    tp_atr_mult: float = 0.9
    sl_atr_mult: float = 0.9
    monitor_candles: int = 3
    candles_buffer: int = 1000
    min_hold_sec: int = 30
    logfile: str = "trades_log.xlsx"
    risk_pct: float = 0.01        # risk per trade (1% of balance)
    margin_type: str = "ISOLATED"
    use_testnet: bool = False     # set True to use Binance Futures testnet


# ---------------- Excel Logger ----------------
def init_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append([
            "pair", "entry_time", "side", "entry_price", "qty", "tp_price", "sl_price",
            "exit_time", "exit_price", "pnl", "fees", "exit_reason", "balance_after"
        ])
        wb.save(path)


def append_trade_excel(path: str, row: List):
    wb = load_workbook(path)
    ws = wb["Trades"]
    clean_row = [v.replace(tzinfo=None) if isinstance(v, datetime) else v for v in row]
    ws.append(clean_row)
    wb.save(path)


# ---------------- Technical helpers ----------------
def compute_atr_from_df(df: pd.DataFrame, period: int):
    high_low = df['high'] - df['low']
    high_close = (df['high'] - df['close'].shift(1)).abs()
    low_close = (df['low'] - df['close'].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    return atr


def compute_qty_by_risk(balance: float, risk_pct: float, entry_price: float, sl_price: float):
    risk_amount = balance * risk_pct
    risk_per_unit = abs(entry_price - sl_price)
    if risk_per_unit == 0:
        raise ValueError("SL tidak boleh sama dengan entry_price")
    qty = risk_amount / risk_per_unit
    return qty


async def round_qty_to_step(client: AsyncClient, symbol: str, raw_qty: float):
    info = await client.futures_exchange_info()
    stepSize = None
    minQty = None
    for s in info['symbols']:
        if s['symbol'] == symbol:
            lot = next((f for f in s['filters'] if f['filterType'] == 'LOT_SIZE'), None)
            if lot:
                stepSize = float(lot['stepSize'])
                minQty = float(lot['minQty'])
            break
    if stepSize is None:
        raise RuntimeError("LOT_SIZE not found for symbol")
    # round down
    rounded = math.floor(raw_qty / stepSize) * stepSize
    if rounded < (minQty or 0):
        return 0.0
    return float(f"{rounded:.8f}")


# ---------------- Live Bot ----------------
class LiveDualEntryBot:
    def __init__(self, cfg: BotConfig):
        self.cfg = cfg
        self.balance = cfg.initial_balance
        self.candles = pd.DataFrame(columns=['open', 'high', 'low', 'close', 'volume'])
        self._current_position: Optional[Dict] = None
        self.watches: List[Dict] = []
        init_excel(self.cfg.logfile)
        self.client: Optional[AsyncClient] = None
        self.bm: Optional[BinanceSocketManager] = None

    async def start(self):
        # create client
        self.client = await AsyncClient.create(self.cfg.api_key, self.cfg.api_secret, testnet=self.cfg.use_testnet)
        # try set leverage & margin
        try:
            await self.client.futures_change_leverage(symbol=self.cfg.pair, leverage=int(self.cfg.leverage))
            # margin type change may raise if not supported
            try:
                await self.client.futures_change_margin_type(symbol=self.cfg.pair, marginType=self.cfg.margin_type)
            except Exception:
                pass
        except Exception as e:
            print("[WARN] set leverage/margin:", e)

        self.bm = BinanceSocketManager(self.client)
        print(f"[INFO] Starting live-dual-entry bot for {self.cfg.pair} (testnet={self.cfg.use_testnet})")
        async with self.bm.kline_socket(self.cfg.pair, interval=self.cfg.interval) as stream:
            while True:
                try:
                    res = await stream.recv()
                    k = res.get('k', {})
                    is_closed = k.get('x', False)
                    ts = pd.to_datetime(k.get('t'), unit='ms', utc=True)
                    o, h, l, c, v = map(float, (k.get('o'), k.get('h'), k.get('l'), k.get('c'), k.get('v')))
                    self._append_candle(ts, o, h, l, c, v)

                    if is_closed:
                        atr_series = compute_atr_from_df(self.candles, self.cfg.atr_period)
                        current_atr = atr_series.iat[-1] if len(atr_series) >= self.cfg.atr_period else np.nan

                        if not np.isnan(current_atr) and current_atr >= self.cfg.min_atr:
                            self._create_watch(current_atr)

                        await self._process_watches()
                        await self._process_current_position()
                except Exception as e:
                    print("[ERROR] main loop:", e)
                    await asyncio.sleep(1)

    def _append_candle(self, ts, o, h, l, c, v):
        # use UTC timestamps (pandas index)
        row = pd.DataFrame([[o, h, l, c, v]], index=[ts], columns=['open', 'high', 'low', 'close', 'volume'])
        self.candles = pd.concat([self.candles, row])
        if len(self.candles) > self.cfg.candles_buffer:
            self.candles = self.candles.iloc[-self.cfg.candles_buffer:]

    def _create_watch(self, atr_value):
        if self._current_position is not None:
            return
        last_close = self.candles['close'].iat[-1]
        watch = {
            "start_idx": len(self.candles) - 1,
            "expire_idx": len(self.candles) - 1 + self.cfg.monitor_candles,
            "long_level": last_close + atr_value * self.cfg.level_mult,
            "short_level": last_close - atr_value * self.cfg.level_mult,
            "atr": atr_value,
            "trigger_time": self.candles.index[-1]
        }
        self.watches.append(watch)
        print(f"[WATCH CREATED] {watch['trigger_time']} ATR={atr_value:.6f} long={watch['long_level']:.6f} short={watch['short_level']:.6f}")

    async def _process_watches(self):
        if self._current_position is not None:
            return

        latest_idx = len(self.candles) - 1
        candle_high = self.candles['high'].iat[-1]
        candle_low = self.candles['low'].iat[-1]

        new_watches = []
        for w in self.watches:
            if latest_idx <= w['start_idx']:
                new_watches.append(w)
                continue

            triggered = False
            side = None
            entry_price = None
            tp_price = None
            sl_price = None

            if candle_high >= w['long_level']:
                triggered = True
                side = 'LONG'
                entry_price = w['long_level']
                tp_price = entry_price + w['atr'] * self.cfg.tp_atr_mult
                sl_price = entry_price - w['atr'] * self.cfg.sl_atr_mult
            elif candle_low <= w['short_level']:
                triggered = True
                side = 'SHORT'
                entry_price = w['short_level']
                tp_price = entry_price - w['atr'] * self.cfg.tp_atr_mult
                sl_price = entry_price + w['atr'] * self.cfg.sl_atr_mult

            if triggered:
                print(f"[TRIGGER] {w['trigger_time']} side={side} entry={entry_price:.6f} tp={tp_price:.6f} sl={sl_price:.6f}")
                await self._open_position(side, entry_price, tp_price, sl_price, w['atr'])
            else:
                if latest_idx < w['expire_idx']:
                    new_watches.append(w)

        self.watches = new_watches

    async def _open_position(self, side: str, entry_level_price: float, tp_price: float, sl_price: float, atr_value: float):
        # compute qty by risk, rounding to step
        raw_qty = compute_qty_by_risk(self.balance, self.cfg.risk_pct, entry_level_price, sl_price)
        qty = await round_qty_to_step(self.client, self.cfg.pair, raw_qty)
        if qty <= 0:
            print("[OPEN SKIP] calculated qty <= 0 (minQty or too small).")
            return

        try:
            # place market order
            order = await self.client.futures_create_order(
                symbol=self.cfg.pair,
                side=SIDE_BUY if side == 'LONG' else SIDE_SELL,
                type=ORDER_TYPE_MARKET,
                quantity=qty
            )
            # get executed price and qty (try multiple fields)
            exec_qty = 0.0
            exec_price = None
            # many responses include 'fills' list
            if isinstance(order, dict):
                if 'fills' in order and order['fills']:
                    # calculate average price weighted by qty
                    num = 0.0
                    den = 0.0
                    for f in order['fills']:
                        fq = float(f.get('qty', f.get('qty', 0)))
                        fp = float(f.get('price', f.get('price', 0)))
                        num += fq * fp
                        den += fq
                    if den > 0:
                        exec_price = num / den
                        exec_qty = den
                # fallback fields
                exec_qty = exec_qty or float(order.get('executedQty', 0) or order.get('origQty', 0))
                if exec_price is None:
                    # some responses include 'avgPrice'
                    try:
                        exec_price = float(order.get('avgPrice')) if order.get('avgPrice') else None
                    except Exception:
                        exec_price = None

            # if still unknown, fallback to last candle close
            if not exec_price or exec_price == 0:
                exec_price = float(self.candles['close'].iat[-1])
            if not exec_qty or exec_qty == 0:
                exec_qty = qty

            print(f"[OPENED] {side} {exec_qty} @ {exec_price:.6f}")

            self._current_position = {
                "side": side,
                "entry_price": exec_price,
                "qty": float(exec_qty),
                "tp_price": tp_price,
                "sl_price": sl_price,
                "entry_time": datetime.now(timezone.utc),
                "atr": atr_value
            }
        except Exception as e:
            print("[ERROR] open position:", e)

    async def _process_current_position(self):
        if self._current_position is None:
            return

        pos = self._current_position
        latest_candle = self.candles.iloc[-1]
        now = self.candles.index[-1]
        elapsed_sec = (now.to_pydatetime() - pos['entry_time']).total_seconds()

        exit_reason = None
        exit_price = None

        if elapsed_sec >= self.cfg.min_hold_sec:
            if pos['side'] == 'LONG':
                if latest_candle['high'] >= pos['tp_price']:
                    exit_price = pos['tp_price']
                    exit_reason = 'TP'
                elif latest_candle['low'] <= pos['sl_price']:
                    exit_price = pos['sl_price']
                    exit_reason = 'SL'
            else:
                if latest_candle['low'] <= pos['tp_price']:
                    exit_price = pos['tp_price']
                    exit_reason = 'TP'
                elif latest_candle['high'] >= pos['sl_price']:
                    exit_price = pos['sl_price']
                    exit_reason = 'SL'

        if exit_price is not None:
            # execute market close (reduceOnly)
            try:
                side = SIDE_SELL if pos['side'] == 'LONG' else SIDE_BUY
                close_order = await self.client.futures_create_order(
                    symbol=self.cfg.pair,
                    side=side,
                    type=ORDER_TYPE_MARKET,
                    reduceOnly=True,
                    quantity=pos['qty']
                )
                # determine executed exit price
                exit_exec_price = None
                exit_exec_qty = 0.0
                if isinstance(close_order, dict):
                    if 'fills' in close_order and close_order['fills']:
                        num = 0.0
                        den = 0.0
                        for f in close_order['fills']:
                            fq = float(f.get('qty', 0))
                            fp = float(f.get('price', 0))
                            num += fq * fp
                            den += fq
                        if den > 0:
                            exit_exec_price = num / den
                            exit_exec_qty = den
                    exit_exec_qty = exit_exec_qty or float(close_order.get('executedQty', 0) or close_order.get('origQty', 0))
                    try:
                        exit_exec_price = exit_exec_price or (float(close_order.get('avgPrice')) if close_order.get('avgPrice') else None)
                    except Exception:
                        pass

                if not exit_exec_price:
                    exit_exec_price = float(latest_candle['close'])

                # compute pnl in USDT: qty * price_diff
                if pos['side'] == 'LONG':
                    raw_pnl = pos['qty'] * (exit_exec_price - pos['entry_price'])
                else:
                    raw_pnl = pos['qty'] * (pos['entry_price'] - exit_exec_price)

                # fees: entry_notional + exit_notional times fee_rate
                entry_notional = pos['qty'] * pos['entry_price']
                exit_notional = pos['qty'] * exit_exec_price
                fees = (entry_notional + exit_notional) * self.cfg.fee_rate

                net_pnl = raw_pnl - fees

                # update balance and log
                self.balance += net_pnl

                append_trade_excel(self.cfg.logfile, [
                    self.cfg.pair,
                    pos['entry_time'],
                    pos['side'],
                    pos['entry_price'],
                    pos['qty'],
                    pos['tp_price'],
                    pos['sl_price'],
                    datetime.now(timezone.utc),
                    exit_exec_price,
                    net_pnl,
                    fees,
                    exit_reason,
                    self.balance
                ])

                print(f"[CLOSED] {pos['side']} exit={exit_exec_price:.6f} reason={exit_reason} pnl={net_pnl:.6f} fees={fees:.6f} balance={self.balance:.4f}")
                self._current_position = None
            except Exception as e:
                print("[ERROR] closing position:", e)


# ---------------- Run ----------------
if __name__ == "__main__":
    cfg = BotConfig()
    # --- IMPORTANT: fill your API keys here BEFORE running ---
    cfg.api_key = os.getenv('BINANCE_API_KEY')    # <-- put API key
    cfg.api_secret = os.getenv('BINANCE_API_SECRET') # <-- put API secret
    cfg.use_testnet = False   # set True if you want testnet

    init_excel(cfg.logfile)
    bot = LiveDualEntryBot(cfg)

    # run event loop
    loop = asyncio.get_event_loop()
    try:
        loop.run_until_complete(bot.start())
    except KeyboardInterrupt:
        print("[STOP] Interrupted by user. Trades saved to", cfg.logfile)
    finally:
        try:
            if bot.client:
                loop.run_until_complete(bot.client.close_connection())
        except Exception:
            pass
