# live_bot_dual_entry_liveclose_improved.py
import asyncio
import math
import os
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from binance import AsyncClient, BinanceSocketManager
from binance.enums import SIDE_BUY, SIDE_SELL, ORDER_TYPE_MARKET, ORDER_TYPE_LIMIT, TIME_IN_FORCE_GTC
from dotenv import load_dotenv
load_dotenv()

# ---------------- Config ----------------
@dataclass
class BotConfig:
    api_key: str = os.getenv('BINANCE_API_KEY')
    api_secret: str = os.getenv('BINANCE_API_SECRET')
    pair: str = "AVAXUSDT"
    interval: str = "5m"
    initial_balance: float = 20.0
    leverage: float = 10.0
    fee_rate: float = 0.0004
    min_atr: float = 0.0005
    atr_period: int = 14
    level_mult: float = 0.15
    tp_atr_mult: float = 0.8
    sl_atr_mult: float = 1.0
    monitor_candles: int = 3
    candles_buffer: int = 1000
    min_hold_sec: int = 900
    logfile: str = "trades_log.xlsx"
    risk_pct: float = 0.008
    margin_type: str = "ISOLATED"
    use_testnet: bool = True
    use_limit_orders: bool = True
    slippage_pct: float = 0.001
    require_confirmation: bool = True
    adaptive_tp_sl: bool = True
    # AVAX-specific precision settings
    qty_precision: int = 1  # 1 decimal for AVAX
    price_precision: int = 3  # 3 decimals for AVAX
    

# ---------------- Excel Logger ----------------
def init_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append([
            "pair", "entry_time", "side", "entry_price", "qty", "tp_price", "sl_price",
            "exit_time", "exit_price", "pnl", "fees", "exit_reason", "balance_after",
            "atr_value", "volatility_ratio"
        ])
        wb.save(path)

def append_trade_excel(path: str, row: List):
    wb = load_workbook(path)
    ws = wb["Trades"]
    clean_row = [v.replace(tzinfo=None) if isinstance(v, datetime) else v for v in row]
    ws.append(clean_row)
    wb.save(path)

# ---------------- Technical helpers ----------------
def compute_atr_from_df(df: pd.DataFrame, period: int):
    high_low = df['high'] - df['low']
    high_close = (df['high'] - df['close'].shift(1)).abs()
    low_close = (df['low'] - df['close'].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    return atr

def compute_volatility_ratio(df: pd.DataFrame, atr_period: int = 14):
    atr = compute_atr_from_df(df, atr_period)
    current_atr = atr.iat[-1] if len(atr) > 0 else 0
    price = df['close'].iat[-1] if len(df) > 0 else 1
    return current_atr / price if price > 0 else 0

def compute_qty_by_risk(balance: float, risk_pct: float, entry_price: float, sl_price: float):
    risk_amount = balance * risk_pct
    risk_per_unit = abs(entry_price - sl_price)
    if risk_per_unit == 0:
        return 0
    qty = risk_amount / risk_per_unit
    return qty


def calculate_order_quality(self, order: Dict) -> float:
    """Calculate a quality score for order prioritization"""
    # Factors to consider:
    # 1. Distance from current price (closer is better)
    # 2. Volatility multiplier (higher volatility might be better for scalping)
    # 3. Time since signal (newer signals might be better)
    
    current_price = self.candles['close'].iloc[-1] if not self.candles.empty else 0
    price_distance = abs(order['entry_price'] - current_price) / current_price if current_price else 1
    
    # Lower distance = higher quality
    distance_score = 1 - min(price_distance, 0.1) * 10  # Normalize to 0-1 range
    
    # Higher volatility multiplier = higher quality (for aggressive trading)
    volatility_score = order.get('vol_mult', 1.0)
    
    # Newer orders are generally better
    time_score = 1.0  # Can be based on order time if needed
    
    return distance_score * 0.5 + volatility_score * 0.3 + time_score * 0.2


async def round_qty_to_step(client: AsyncClient, symbol: str, raw_qty: float):
    info = await client.futures_exchange_info()
    stepSize = None
    minQty = None
    for s in info['symbols']:
        if s['symbol'] == symbol:
            lot = next((f for f in s['filters'] if f['filterType'] == 'LOT_SIZE'), None)
            if lot:
                stepSize = float(lot['stepSize'])
                minQty = float(lot['minQty'])
            break
    if stepSize is None:
        raise RuntimeError("LOT_SIZE not found for symbol")
    rounded = math.floor(raw_qty / stepSize) * stepSize
    if rounded < (minQty or 0):
        return 0.0
    return float(f"{rounded:.8f}")

# ---------------- Live Bot ----------------
class ImprovedLiveDualEntryBot:
    def __init__(self, cfg: BotConfig):
        self.cfg = cfg
        self.balance = cfg.initial_balance
        self.candles = pd.DataFrame(columns=['open', 'high', 'low', 'close', 'volume'])
        self._current_position: Optional[Dict] = None
        self.watches: List[Dict] = []
        self.pending_orders: List[Dict] = []
        init_excel(self.cfg.logfile)
        self.client: Optional[AsyncClient] = None
        self.bm: Optional[BinanceSocketManager] = None
        self.volatility_ratio = 0.0
        self.pending_orders: List[Dict] = []  # Sudah ada
        self.active_orders: List[Dict] = []   # Order yang sudah aktif tapi belum ditutup
        self.order_timeout = 300  # 5 menit timeout untuk pending orders
    
    async def check_order_timeouts(self):
        """Cancel orders that have been pending too long"""
        now = datetime.now(timezone.utc)
        for order in self.pending_orders[:]:
            if (now - order['place_time']).total_seconds() > self.order_timeout:
                try:
                    await self.client.futures_cancel_order(
                        symbol=self.cfg.pair,
                        orderId=order['order_id']
                    )
                    self.pending_orders.remove(order)
                    print(f"[ORDER CANCELED] Timeout - {order['order_id']}")
                except Exception as e:
                    print(f"[ERROR] canceling timed out order: {e}")

    def compute_ema(series: pd.Series, period: int):
        return series.ewm(span=period, adjust=False).mean()

    def compute_adx(df: pd.DataFrame, period: int = 14):
        high = df['high']
        low = df['low']
        close = df['close']

        prev_close = close.shift(1)
        tr1 = high - low
        tr2 = (high - prev_close).abs()
        tr3 = (low - prev_close).abs()
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)

        up_move = high.diff()
        down_move = -low.diff()
        plus_dm = ((up_move > down_move) & (up_move > 0)) * up_move
        minus_dm = ((down_move > up_move) & (down_move > 0)) * down_move

        atr = tr.rolling(window=period).mean()
        plus_di = 100 * (plus_dm.rolling(window=period).sum() / atr)
        minus_di = 100 * (minus_dm.rolling(window=period).sum() / atr)

        dx = ((plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, np.nan)) * 100
        adx = dx.rolling(window=period).mean()
        return adx.fillna(0)

    async def manage_order_conflicts(self, new_order: Dict):
        """Handle multiple order requests based on priority"""
        # Strategy 1: Cancel oldest order if new order has better potential
        if len(self.pending_orders) >= 2:  # Max 2 pending orders
            # Cancel the oldest order
            oldest_order = min(self.pending_orders, key=lambda x: x['place_time'])
            try:
                await self.client.futures_cancel_order(
                    symbol=self.cfg.pair,
                    orderId=oldest_order['order_id']
                )
                self.pending_orders.remove(oldest_order)
                print(f"[ORDER CANCELED] Conflict resolution - removed oldest order {oldest_order['order_id']}")
            except Exception as e:
                print(f"[ERROR] canceling order: {e}")
        
        # Strategy 2: Compare order quality and keep the best one
        current_best = None
        if self.pending_orders:
            current_best = max(self.pending_orders, key=lambda x: x.get('quality_score', 0))
        
        new_order_quality = self.calculate_order_quality(new_order)
        
        if current_best and new_order_quality > current_best.get('quality_score', 0):
            # New order is better, cancel the current one
            try:
                await self.client.futures_cancel_order(
                    symbol=self.cfg.pair,
                    orderId=current_best['order_id']
                )
                self.pending_orders.remove(current_best)
                print(f"[ORDER CANCELED] Replaced with better order")
            except Exception as e:
                print(f"[ERROR] canceling order: {e}")

    def calculate_order_quality(self, order: Dict) -> float:
        """Calculate a quality score for order prioritization"""
        # Factors to consider:
        # 1. Distance from current price (closer is better)
        # 2. Volatility multiplier (higher volatility might be better for scalping)
        # 3. Time since signal (newer signals might be better)
        
        current_price = self.candles['close'].iloc[-1] if not self.candles.empty else 0
        price_distance = abs(order['entry_price'] - current_price) / current_price if current_price else 1
        
        # Lower distance = higher quality
        distance_score = 1 - min(price_distance, 0.1) * 10  # Normalize to 0-1 range
        
        # Higher volatility multiplier = higher quality (for aggressive trading)
        volatility_score = order.get('vol_mult', 1.0)
        
        # Newer orders are generally better
        time_score = 1.0  # Can be based on order time if needed
        
        return distance_score * 0.5 + volatility_score * 0.3 + time_score * 0.2
    
    async def check_order_status(self):
        """Periodically check status of all pending orders"""
        if not self.client:
            return
            
        for order in self.pending_orders[:]:
            try:
                # Check order status from exchange
                order_status = await self.client.futures_get_order(
                    symbol=self.cfg.pair,
                    orderId=order['order_id']
                )
                
                if order_status['status'] == 'FILLED':
                    # Order filled, move to active orders
                    print(f"[ORDER FILLED] {order['side']} {order['qty']} @ {order_status['avgPrice']}")
                    self.pending_orders.remove(order)
                    self.active_orders.append({
                        **order,
                        'entry_price': float(order_status['avgPrice']),
                        'entry_time': datetime.now(timezone.utc),
                        'status': 'FILLED'
                    })

                    self._current_position = {
                        **order,
                        'entry_price': float(order_status['avgPrice']),
                        'entry_time': datetime.now(timezone.utc),
                        'status': 'FILLED'
                    }

                elif order_status['status'] == 'CANCELED' or order_status['status'] == 'EXPIRED':
                    # Order canceled or expired, remove from pending
                    print(f"[ORDER {order_status['status']}] {order['order_id']}")
                    self.pending_orders.remove(order)
                    
            except Exception as e:
                print(f"[ERROR] checking order status: {e}")

    def _round_price(self, price: float) -> float:
        """Round price to configured precision"""
        return round(price, self.cfg.price_precision)

    def _round_qty(self, qty: float) -> float:
        """Round quantity to configured precision"""
        return round(qty, self.cfg.qty_precision)

    async def periodic_order_checks(self):
        """Run order checks periodically"""
        while True:
            try:
                await asyncio.sleep(30)  # Check every 30 seconds
                await self.check_order_status()
                await self.check_order_timeouts()
            except Exception as e:
                print("[ERROR] in periodic order checks:", e)
    
    async def _close_position(self, side: str, exit_price: float, reason: str = "EMERGENCY"):
        if self._current_position is None:
            return
        
        pos = self._current_position
        try:
            # market close order
            close_order = await self.client.futures_create_order(
                symbol=self.cfg.pair,
                side=SIDE_SELL if side == 'LONG' else SIDE_BUY,
                type=ORDER_TYPE_MARKET,
                reduceOnly=True,
                quantity=pos['qty']
            )

            # Hitung harga rata-rata eksekusi
            exit_exec_price = 0.0
            exit_qty = 0.0
            if 'fills' in close_order and close_order['fills']:
                for fill in close_order['fills']:
                    exit_qty += float(fill['qty'])
                    exit_exec_price += float(fill['price']) * float(fill['qty'])
                exit_exec_price = exit_exec_price / exit_qty if exit_qty > 0 else exit_price
            else:
                exit_exec_price = exit_price

            # Sesuaikan slippage
            if pos['side'] == 'LONG':
                exit_exec_price *= 1 - self.cfg.slippage_pct
            else:
                exit_exec_price *= 1 + self.cfg.slippage_pct

            exit_exec_price = self._round_price(exit_exec_price)

            # Hitung PnL
            raw_pnl = pos['qty'] * ((exit_exec_price - pos['entry_price']) if pos['side'] == 'LONG' else (pos['entry_price'] - exit_exec_price))
            fees = self.cfg.fee_rate * self.balance
            net_pnl = raw_pnl - fees
            self.balance += net_pnl

            self._current_position = None
            # self.watches.clear()

            # Log ke Excel
            append_trade_excel(self.cfg.logfile, [
                self.cfg.pair,
                pos['entry_time'],
                pos['side'],
                pos['entry_price'],
                pos['qty'],
                pos['tp_price'],
                pos['sl_price'],
                datetime.now(timezone.utc),
                exit_exec_price,
                net_pnl,
                fees,
                reason,
                self.balance,
                pos['atr'],
                pos['vol_mult']
            ])

            print(f"[POSITION CLOSED] {pos['side']} exit={exit_exec_price:.3f} reason={reason} pnl={net_pnl:.6f} balance={self.balance:.4f}")
            
        except Exception as e:
            print("[ERROR] closing position:", e)

    
    async def _emergency_exit_check(self, price: float):

        if self._current_position is None:
            return

        pos = self._current_position
        side = pos['side']
        entry_price = pos['entry_price']
        tp = pos['tp_price']
        sl = pos['sl_price']

        # LONG position
        if side == "LONG":
            if price <= sl:
                print(f"[EMERGENCY EXIT] LONG SL hit @ {price}")
                await self._close_position("LONG", price, reason="EMERGENCY")
                self._current_position = None
                # self.watches.clear()
            elif price >= tp:
                print(f"[EMERGENCY EXIT] LONG TP hit @ {price}")
                await self._close_position("LONG", price, reason="EMERGENCY")
                self._current_position = None
                # self.watches.clear()

        # SHORT position
        elif side == "SHORT":
            if price >= sl:
                print(f"[EMERGENCY EXIT] SHORT SL hit @ {price}")
                await self._close_position("SHORT", price, reason="EMERGENCY")
                self._current_position = None
                # self.watches.clear()
            elif price <= tp:
                print(f"[EMERGENCY EXIT] SHORT TP hit @ {price}")
                await self._close_position("SHORT", price, reason="EMERGENCY")
                self._current_position = None
                # self.watches.clear()
                self.watches = [w for w in self.watches if w['expire_idx'] > len(self.candles)-1]

                
    async def start(self):
        self.client = await AsyncClient.create(self.cfg.api_key, self.cfg.api_secret, testnet=self.cfg.use_testnet)
        try:
            await self.client.futures_change_leverage(symbol=self.cfg.pair, leverage=int(self.cfg.leverage))
            try:
                await self.client.futures_change_margin_type(symbol=self.cfg.pair, marginType=self.cfg.margin_type)
            except Exception:
                pass
        except Exception as e:
            print("[WARN] set leverage/margin:", e)
        
        

        self.bm = BinanceSocketManager(self.client)
        print(f"[INFO] Starting improved live-dual-entry bot for {self.cfg.pair} (testnet={self.cfg.use_testnet})")
        print(f"[CONFIG] Price precision: {self.cfg.price_precision}, Qty precision: {self.cfg.qty_precision}")
        
        asyncio.create_task(self.periodic_order_checks())
        async with self.bm.kline_futures_socket(self.cfg.pair, interval=self.cfg.interval) as stream:
            while True:
                try:
                    res = await stream.recv()
                    k = res.get('k', {})
                    is_closed = k.get('x', False)
                    ts = pd.to_datetime(k.get('t'), unit='ms', utc=True)
                    o, h, l, c, v = map(float, (k.get('o'), k.get('h'), k.get('l'), k.get('c'), k.get('v')))
                    self._append_candle(ts, o, h, l, c, v)

                    self.volatility_ratio = compute_volatility_ratio(self.candles, self.cfg.atr_period)
                    last_price = float(k.get('c', 0))

                    await self._emergency_exit_check(last_price)
                    if is_closed:
                        atr_series = compute_atr_from_df(self.candles, self.cfg.atr_period)
                        current_atr = atr_series.iat[-1] if len(atr_series) >= self.cfg.atr_period else np.nan

                        if not np.isnan(current_atr) and current_atr >= self.cfg.min_atr:
                            self._create_watch(current_atr)

                        await self._process_watches()
                        await self._process_current_position()
                        await self._check_pending_orders()
                        await self.check_order_status()
                        await self.check_order_timeouts()

                except Exception as e:
                    print("[ERROR] main loop:", e)
                    await asyncio.sleep(1)


    def _append_candle(self, ts, o, h, l, c, v):
        row = pd.DataFrame([[o, h, l, c, v]], index=[ts], columns=['open', 'high', 'low', 'close', 'volume'])
        if self.candles.empty:
            self.candles = row
        else:
            self.candles = pd.concat([self.candles, row])
        if len(self.candles) > self.cfg.candles_buffer:
            self.candles = self.candles.iloc[-self.cfg.candles_buffer:]

    def _create_watch(self, atr_value):
        if self._current_position is not None:
            return
        last_close = self.candles['close'].iat[-1]
        
        volatility_mult = 1.0
        if self.cfg.adaptive_tp_sl and self.volatility_ratio > 0.005:
            volatility_mult = 1.2
        elif self.cfg.adaptive_tp_sl and self.volatility_ratio < 0.002:
            volatility_mult = 0.8
            
        watch = {
            "start_idx": len(self.candles) - 1,
            "expire_idx": len(self.candles) - 1 + self.cfg.monitor_candles,
            "long_level": self._round_price(last_close + atr_value * self.cfg.level_mult * volatility_mult),
            "short_level": self._round_price(last_close - atr_value * self.cfg.level_mult * volatility_mult),
            "atr": atr_value,
            "trigger_time": self.candles.index[-1],
            "volatility_mult": volatility_mult
        }
        self.watches.append(watch)
        print(f"[WATCH CREATED] {watch['trigger_time']} ATR={atr_value:.6f} long={watch['long_level']:.3f} short={watch['short_level']:.3f} vol_mult={volatility_mult:.2f}")

    async def _process_watches(self):
        print("[INFO] Masuk Processing watches...")
        if self._current_position is not None:
            return
        print("[INFO] Mulai Processing watches...")
        latest_idx = len(self.candles) - 1
        candle_high = self.candles['high'].iat[-1]
        candle_low = self.candles['low'].iat[-1]
        candle_close = self.candles['close'].iat[-1]

        new_watches = []
        for w in self.watches:
            if latest_idx <= w['start_idx']:
                new_watches.append(w)
                continue

            triggered = False
            side = None
            entry_price = None
            tp_price = None
            sl_price = None

            if self.cfg.require_confirmation:
                long_condition = candle_close >= w['long_level']
                short_condition = candle_close <= w['short_level']
            else:
                long_condition = candle_high >= w['long_level']
                short_condition = candle_low <= w['short_level']

            print(f"long_condition: {long_condition}, short_condition: {short_condition}")

            if long_condition:
                triggered = True
                side = 'LONG'
                entry_price = w['long_level']
                tp_mult = self.cfg.tp_atr_mult * w['volatility_mult']
                sl_mult = self.cfg.sl_atr_mult * w['volatility_mult']
                tp_price = self._round_price(entry_price + w['atr'] * tp_mult)
                sl_price = self._round_price(entry_price - w['atr'] * sl_mult)
            elif short_condition:
                triggered = True
                side = 'SHORT'
                entry_price = w['short_level']
                tp_mult = self.cfg.tp_atr_mult * w['volatility_mult']
                sl_mult = self.cfg.sl_atr_mult * w['volatility_mult']
                tp_price = self._round_price(entry_price - w['atr'] * tp_mult)
                sl_price = self._round_price(entry_price + w['atr'] * sl_mult)

            if triggered:
                print(f"[TRIGGER] {w['trigger_time']} side={side} entry={entry_price:.3f} tp={tp_price:.3f} sl={sl_price:.3f}")
                if self.cfg.use_limit_orders:
                    await self._place_limit_order(side, entry_price, tp_price, sl_price, w['atr'], w['volatility_mult'])
                else:
                    await self._open_market_position(side, entry_price, tp_price, sl_price, w['atr'], w['volatility_mult'])
            else:
                if latest_idx < w['expire_idx']:
                    new_watches.append(w)
        

        self.watches = new_watches

    async def _place_limit_order(self, side: str, entry_price: float, tp_price: float, sl_price: float, atr_value: float, vol_mult: float):
        qty = self._round_qty(1.0)  # 1 AVAX

        if side == 'LONG':
            order_price = entry_price * (1 - 0.0002)  # 0.05% below trigger
        else:
            order_price = entry_price * (1 + 0.0002)  # 0.05% above trigger

        # Round prices
        order_price = self._round_price(order_price)
        entry_price = self._round_price(entry_price)
        tp_price = self._round_price(tp_price)
        sl_price = self._round_price(sl_price)

        risk_per_trade = abs(entry_price - sl_price) * qty
        risk_percentage = (risk_per_trade / self.balance) * 100
        print(f"[DEBUG] Actual risk: {risk_percentage:.2f}% per trade")

        if len(self.pending_orders) >= 2:  # Maximum 2 pending orders
            await self.manage_order_conflicts({
                'side': side,
                'entry_price': entry_price,
                'tp_price': tp_price,
                'sl_price': sl_price,
                'atr': atr_value,
                'vol_mult': vol_mult,
                'qty': qty
            })
        
        try:
            # Place the order
            order = await self.client.futures_create_order(
                symbol=self.cfg.pair,
                side=SIDE_BUY if side == 'LONG' else SIDE_SELL,
                type=ORDER_TYPE_LIMIT,
                timeInForce=TIME_IN_FORCE_GTC,
                quantity=qty,
                price=order_price
            )
            
            # Add to pending orders with additional metadata
            order_data = {
                "order_id": order['orderId'],
                "side": side,
                "entry_price": entry_price,
                "order_price": order_price,
                "qty": qty,
                "tp_price": tp_price,
                "sl_price": sl_price,
                "atr": atr_value,
                "vol_mult": vol_mult,
                "place_time": datetime.now(timezone.utc),
                "expiry_time": datetime.now(timezone.utc) + timedelta(seconds=self.order_timeout),
                "quality_score": self.calculate_order_quality({
                    'entry_price': entry_price,
                    'vol_mult': vol_mult
                })
            }
            
            self.pending_orders.append(order_data)
            print(f"[LIMIT ORDER PLACED] {side} {qty} @ {order_price:.3f} (orderId: {order['orderId']})")
            
        except Exception as e:
            print("[ERROR] place limit order:", e)

        # try:
        #     order = await self.client.futures_create_order(
        #         symbol=self.cfg.pair,
        #         side=SIDE_BUY if side == 'LONG' else SIDE_SELL,
        #         type=ORDER_TYPE_LIMIT,
        #         timeInForce=TIME_IN_FORCE_GTC,
        #         quantity=qty,
        #         price=order_price
        #     )
            
        #     self.pending_orders.append({
        #         "order_id": order['orderId'],
        #         "side": side,
        #         "entry_price": entry_price,
        #         "order_price": order_price,
        #         "qty": qty,
        #         "tp_price": tp_price,
        #         "sl_price": sl_price,
        #         "atr": atr_value,
        #         "vol_mult": vol_mult,
        #         "place_time": datetime.now(timezone.utc),
        #         "expiry_time": datetime.now(timezone.utc) + timedelta(minutes=5)
        #     })
        #     print(f"[LIMIT ORDER PLACED] {side} {qty} @ {order_price:.3f} (orderId: {order['orderId']})")
        # except Exception as e:
        #     print("[ERROR] place limit order:", e)

    async def _open_market_position(self, side: str, entry_price: float, tp_price: float, sl_price: float, atr_value: float, vol_mult: float):
        qty = self._round_qty(1.0)  # 1 AVAX

        # Round prices
        entry_price = self._round_price(entry_price)
        tp_price = self._round_price(tp_price)
        sl_price = self._round_price(sl_price)

        risk_per_trade = abs(entry_price - sl_price) * qty
        risk_percentage = (risk_per_trade / self.balance) * 100
        print(f"[DEBUG] Actual risk: {risk_percentage:.2f}% per trade")

        try:
            if side == 'LONG':
                exec_price = entry_price * (1 + self.cfg.slippage_pct)
            else:
                exec_price = entry_price * (1 - self.cfg.slippage_pct)
                
            order = await self.client.futures_create_order(
                symbol=self.cfg.pair,
                side=SIDE_BUY if side == 'LONG' else SIDE_SELL,
                type=ORDER_TYPE_MARKET,
                quantity=qty
            )
            
            exec_qty = 0.0
            avg_price = 0.0
            if 'fills' in order and order['fills']:
                for fill in order['fills']:
                    exec_qty += float(fill['qty'])
                    avg_price += float(fill['price']) * float(fill['qty'])
                avg_price = avg_price / exec_qty if exec_qty > 0 else exec_price
            else:
                avg_price = exec_price
                exec_qty = qty

            avg_price = self._round_price(avg_price)
            print(f"[POSITION OPENED] {side} {exec_qty} @ {avg_price:.3f}")

            self._current_position = {
                "side": side,
                "entry_price": avg_price,
                "qty": float(exec_qty),
                "tp_price": tp_price,
                "sl_price": sl_price,
                "entry_time": datetime.now(timezone.utc),
                "atr": atr_value,
                "vol_mult": vol_mult
            }
        except Exception as e:
            print("[ERROR] open position:", e)

    async def _check_pending_orders(self):
        now = datetime.now(timezone.utc)
        for order in self.pending_orders[:]:
            if now >= order['expiry_time']:
                try:
                    await self.client.futures_cancel_order(
                        symbol=self.cfg.pair, 
                        orderId=order['order_id']
                    )
                    print(f"[ORDER CANCELED] {order['order_id']} (expired)")
                    self.pending_orders.remove(order)
                except Exception as e:
                    print(f"[ERROR] cancel order {order['order_id']}:", e)

    async def _process_current_position(self):
        if self._current_position is None:
            return

        pos = self._current_position
        latest_candle = self.candles.iloc[-1]
        now = self.candles.index[-1]
        elapsed_sec = (now.to_pydatetime() - pos['entry_time']).total_seconds()

        exit_reason = None
        exit_price = None
       
        # print('[DEBUG] pos[side]:', pos['side'])
        if elapsed_sec >= self.cfg.min_hold_sec:
            if pos['side'] == 'LONG':
                if latest_candle['high'] >= pos['tp_price']:
                    exit_price = pos['tp_price']
                    exit_reason = 'TP'
                elif latest_candle['low'] <= pos['sl_price']:
                    exit_price = pos['sl_price']
                    exit_reason = 'SL'
                print('exit dari candle')
            else:
                if latest_candle['low'] <= pos['tp_price']:
                    exit_price = pos['tp_price']
                    exit_reason = 'TP'
                elif latest_candle['high'] >= pos['sl_price']:
                    exit_price = pos['sl_price']
                    exit_reason = 'SL'
                print('exit dari candle')

        if exit_price is not None:
            try:
                side = SIDE_SELL if pos['side'] == 'LONG' else SIDE_BUY
                close_order = await self.client.futures_create_order(
                    symbol=self.cfg.pair,
                    side=side,
                    type=ORDER_TYPE_MARKET,
                    reduceOnly=True,
                    quantity=pos['qty']
                )
                
                exit_exec_price = 0.0
                exit_qty = 0.0
                self._current_position = None
                # self.watches.clear()
                if 'fills' in close_order and close_order['fills']:
                    for fill in close_order['fills']:
                        exit_qty += float(fill['qty'])
                        exit_exec_price += float(fill['price']) * float(fill['qty'])
                    exit_exec_price = exit_exec_price / exit_qty if exit_qty > 0 else exit_price
                else:
                    exit_exec_price = exit_price
                    
                if pos['side'] == 'LONG':
                    exit_exec_price = exit_exec_price * (1 - self.cfg.slippage_pct)
                else:
                    exit_exec_price = exit_exec_price * (1 + self.cfg.slippage_pct)

                exit_exec_price = self._round_price(exit_exec_price)

                if pos['side'] == 'LONG':
                    raw_pnl = pos['qty'] * (exit_exec_price - pos['entry_price'])
                else:
                    raw_pnl = pos['qty'] * (pos['entry_price'] - exit_exec_price)

                fees = self.cfg.fee_rate * self.balance
                net_pnl = raw_pnl - fees

                self.balance += net_pnl

                append_trade_excel(self.cfg.logfile, [
                    self.cfg.pair,
                    pos['entry_time'],
                    pos['side'],
                    pos['entry_price'],
                    pos['qty'],
                    pos['tp_price'],
                    pos['sl_price'],
                    datetime.now(timezone.utc),
                    exit_exec_price,
                    net_pnl,
                    fees,
                    exit_reason,
                    self.balance,
                    pos['atr'],
                    pos['vol_mult']
                ])

                print(f"[POSITION CLOSED] {pos['side']} exit={exit_exec_price:.3f} reason={exit_reason} pnl={net_pnl:.6f} fees={fees:.6f} balance={self.balance:.4f}")
                self._current_position = None
                self.watches = [w for w in self.watches if w['expire_idx'] > len(self.candles)-1]
                
            except Exception as e:
                print("[ERROR] closing position:", e)

# ---------------- Run ----------------
if __name__ == "__main__":
    cfg = BotConfig()
    cfg.api_key = os.getenv('BINANCE_API_KEY')
    cfg.api_secret = os.getenv('BINANCE_API_SECRET')
    cfg.use_testnet = False
    
    cfg.use_limit_orders = True
    cfg.require_confirmation = True
    cfg.adaptive_tp_sl = True
    cfg.slippage_pct = 0.001
    cfg.qty_precision = 1
    cfg.price_precision = 3

    init_excel(cfg.logfile)
    bot = ImprovedLiveDualEntryBot(cfg)

    loop = asyncio.get_event_loop()
    try:
        loop.run_until_complete(bot.start())
    except KeyboardInterrupt:
        print("[STOP] Interrupted by user. Trades saved to", cfg.logfile)
    finally:
        try:
            if bot.client:
                loop.run_until_complete(bot.client.close_connection())
        except Exception:
            pass