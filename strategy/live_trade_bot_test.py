# live_bot_dual_entry_v2.py
import asyncio
import os
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Optional

import numpy as np
import pandas as pd
from binance import AsyncClient, BinanceSocketManager
from openpyxl import Workbook, load_workbook

# ---------------- Config ----------------
@dataclass
class BotConfig:
    pair: str = "AVAXUSDT"
    interval: str = "1m"
    initial_balance: float = 20.0
    leverage: float = 20.0
    fee_rate: float = 0.0004
    min_atr: float = 0.0005
    atr_period: int = 14
    level_mult: float = 0.2
    tp_atr_mult: float = 0.9
    sl_atr_mult: float = 0.9
    monitor_candles: int = 3
    candles_buffer: int = 500
    min_hold_sec: int = 30    # minimal holding time sebelum exit
    logfile: str = "trades_log.xlsx"

cfg = BotConfig()

# ---------------- Excel helpers ----------------
def init_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append([
            "pair", "watch_start", "trigger_side", "trigger_level", "atr_at_trigger",
            "entry_time", "entry_price", "exit_time", "exit_price", "pnl",
            "exit_reason", "balance_after"
        ])
        wb.save(path)

def append_trade_excel(path: str, row: List):
    wb = load_workbook(path)
    ws = wb["Trades"]
    clean_row = [v.replace(tzinfo=None) if isinstance(v, datetime) else v for v in row]
    ws.append(clean_row)
    wb.save(path)

# ---------------- Technical helpers ----------------
def compute_atr_from_df(df: pd.DataFrame, period: int):
    high_low = df['high'] - df['low']
    high_close = (df['high'] - df['close'].shift(1)).abs()
    low_close = (df['low'] - df['close'].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    return atr

# ---------------- Live Bot ----------------
class LiveDualEntryBot:
    def __init__(self, cfg: BotConfig):
        self.cfg = cfg
        self.balance = cfg.initial_balance
        self.candles = pd.DataFrame(columns=['open','high','low','close','volume'])
        self._current_position: Optional[Dict] = None
        self.watches: List[Dict] = []
        init_excel(self.cfg.logfile)

    async def start(self):
        client = await AsyncClient.create()
        bm = BinanceSocketManager(client)
        async with bm.kline_socket(self.cfg.pair, interval=self.cfg.interval) as stream:
            print(f"[INFO] Live dual-entry forward-test for {self.cfg.pair}")
            while True:
                res = await stream.recv()
                k = res.get('k', {})
                is_closed = k.get('x', False)
                ts = pd.to_datetime(k.get('t'), unit='ms', utc=True)
                o, h, l, c, v = map(float, (k.get('o'), k.get('h'), k.get('l'), k.get('c'), k.get('v')))
                self._append_candle(ts, o, h, l, c, v)

                if is_closed:
                    atr_series = compute_atr_from_df(self.candles, self.cfg.atr_period)
                    current_atr = atr_series.iat[-1] if len(atr_series) >= self.cfg.atr_period else np.nan

                    if not np.isnan(current_atr) and current_atr >= self.cfg.min_atr:
                        self._create_watch(current_atr)

                    self._process_watches()
                    self._process_current_position()

    def _append_candle(self, ts, o, h, l, c, v):
        self.candles.loc[ts] = [o, h, l, c, v]
        if len(self.candles) > self.cfg.candles_buffer:
            self.candles = self.candles.iloc[-self.cfg.candles_buffer:]

    def _create_watch(self, atr_value):
        if self._current_position is not None:
            # jangan buat watch baru jika posisi masih terbuka
            return
        last_close = self.candles['close'].iat[-1]
        watch = {
            "start_idx": len(self.candles)-1,
            "expire_idx": len(self.candles)-1 + self.cfg.monitor_candles,
            "long_level": last_close + atr_value * self.cfg.level_mult,
            "short_level": last_close - atr_value * self.cfg.level_mult,
            "atr": atr_value,
            "trigger_time": self.candles.index[-1]
        }
        self.watches.append(watch)
        print(f"[WATCH CREATED] {watch['trigger_time']} ATR={atr_value:.6f} long={watch['long_level']:.6f} short={watch['short_level']:.6f}")

    def _process_watches(self):
        if self._current_position is not None:
            # sudah ada posisi terbuka
            return

        latest_idx = len(self.candles) - 1
        candle_high = self.candles['high'].iat[-1]
        candle_low = self.candles['low'].iat[-1]

        new_watches = []
        for w in self.watches:
            if latest_idx <= w['start_idx']:
                new_watches.append(w)
                continue

            triggered = False
            side, entry_price = None, None
            tp_price, sl_price = None, None

            # cek long trigger
            if candle_high >= w['long_level']:
                triggered = True
                side = 'LONG'
                entry_price = w['long_level']
                tp_price = entry_price + w['atr'] * self.cfg.tp_atr_mult
                sl_price = entry_price - w['atr'] * self.cfg.sl_atr_mult
            # cek short trigger
            elif candle_low <= w['short_level']:
                triggered = True
                side = 'SHORT'
                entry_price = w['short_level']
                tp_price = entry_price - w['atr'] * self.cfg.tp_atr_mult
                sl_price = entry_price + w['atr'] * self.cfg.sl_atr_mult

            if triggered:
                self._current_position = {
                    "side": side,
                    "entry_price": entry_price,
                    "tp_price": tp_price,
                    "sl_price": sl_price,
                    "entry_time": w['trigger_time'],
                    "atr": w['atr']
                }
            else:
                if latest_idx < w['expire_idx']:
                    new_watches.append(w)
        self.watches = new_watches

    def _process_current_position(self):
        if self._current_position is None:
            return

        pos = self._current_position
        latest_candle = self.candles.iloc[-1]
        now = self.candles.index[-1]
        elapsed_sec = (now - pos['entry_time']).total_seconds()

        exit_reason = None
        exit_price = None

        # TP/SL hanya diperiksa jika sudah lewat minimal holding time
        if elapsed_sec >= self.cfg.min_hold_sec:
            if pos['side'] == 'LONG':
                if latest_candle['high'] >= pos['tp_price']:
                    exit_price = pos['tp_price']
                    exit_reason = 'TP'
                elif latest_candle['low'] <= pos['sl_price']:
                    exit_price = pos['sl_price']
                    exit_reason = 'SL'
            else:
                if latest_candle['low'] <= pos['tp_price']:
                    exit_price = pos['tp_price']
                    exit_reason = 'TP'
                elif latest_candle['high'] >= pos['sl_price']:
                    exit_price = pos['sl_price']
                    exit_reason = 'SL'

        if exit_price is not None:
            pnl = self._compute_pnl(pos['entry_price'], exit_price, pos['side'])
            self.balance += pnl
            trade = {
                "pair": self.cfg.pair,
                "watch_start": pos['entry_time'],
                "trigger_side": pos['side'],
                "trigger_level": pos['entry_price'],
                "atr_at_trigger": pos['atr'],
                "entry_time": pos['entry_time'],
                "entry_price": pos['entry_price'],
                "exit_time": now,
                "exit_price": exit_price,
                "pnl": pnl,
                "exit_reason": exit_reason,
                "balance_after": self.balance
            }
            append_trade_excel(self.cfg.logfile, [
                trade['pair'], trade['watch_start'], trade['trigger_side'], trade['trigger_level'],
                trade['atr_at_trigger'], trade['entry_time'], trade['entry_price'],
                trade['exit_time'], trade['exit_price'], trade['pnl'],
                trade['exit_reason'], trade['balance_after']
            ])
            print(f"[TRADE] {now} {pos['side']} entry={pos['entry_price']:.6f} exit={exit_price:.6f} pnl={pnl:.6f} balance={self.balance:.4f}")
            self._current_position = None  # posisi sudah ditutup

    def _compute_pnl(self, entry_price: float, exit_price: float, side: str):
        if side == 'LONG':
            frac = (exit_price - entry_price) / entry_price
        else:
            frac = (entry_price - exit_price) / entry_price
        pnl = frac * self.balance * self.cfg.leverage - self.cfg.fee_rate * self.balance
        return pnl

# ---------------- Run ----------------
if __name__ == "__main__":
    init_excel(cfg.logfile)
    bot = LiveDualEntryBot(cfg)
    try:
        asyncio.get_event_loop().run_until_complete(bot.start())
    except KeyboardInterrupt:
        print("[STOP] Interrupted by user. Trades saved to", cfg.logfile)
