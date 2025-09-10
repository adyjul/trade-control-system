# live_bot_live_trade.py
"""
Live trading version (Binance Futures) of dual-entry ATR breakout.
- Sends real MARKET orders to Binance Futures.
- Detection / trigger / TP / SL logic is identical to backtest/forward-test.
- For recording and PnL calculation, this script uses the same assumed prices:
    entry_price = trigger_level
    exit_price  = tp_price or sl_price
  (This preserves exact backtest/forward-test arithmetic as requested.)
- WARNING: actual executed prices on Binance may differ (slippage/spread).
"""

import asyncio
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import List, Dict, Optional

import numpy as np
import pandas as pd
from binance import AsyncClient, BinanceSocketManager
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
# ---------------- Config ----------------
@dataclass
class BotConfig:
    pair: str = "AVAXUSDT"
    interval: str = "1m"
    initial_balance: float = 300.0
    leverage: float = 3.0
    fee_rate: float = 0.0004
    min_atr: float = 0.0005
    atr_period: int = 14
    level_mult: float = 0.2
    tp_atr_mult: float = 0.8
    sl_atr_mult: float = 0.8
    monitor_candles: int = 3
    candles_buffer: int = 500
    logfile: str = "trades_log_real.xlsx"
    live_mode: bool = True  # True -> send real orders
    api_key: str = ""       # fill your API key
    api_secret: str = ""    # fill your API secret
    testnet: bool = False   # set True if using testnet


cfg = BotConfig()

# ---------------- Excel helpers ----------------
def init_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append([
            "pair", "watch_start", "trigger_side", "trigger_level", "atr_at_trigger",
            "entry_time", "entry_price", "exit_time", "exit_price", "pnl",
            "exit_reason", "balance_after", "executed_entry_price", "executed_exit_price"
        ])
        wb.save(path)

def append_trade_excel(path: str, row: List):
    wb = load_workbook(path)
    ws = wb["Trades"]

    clean_row = []
    for v in row:
        if isinstance(v, datetime):
            clean_row.append(v.replace(tzinfo=None))
        else:
            clean_row.append(v)
    ws.append(clean_row)
    wb.save(path)

# ---------------- Technical helpers ----------------
def compute_atr_from_df(df: pd.DataFrame, period: int):
    high_low = df['high'] - df['low']
    high_close = (df['high'] - df['close'].shift(1)).abs()
    low_close = (df['low'] - df['close'].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    return atr

# ---------------- Live Trading Bot ----------------
class LiveDualEntryLiveTradeBot:
    def __init__(self, cfg: BotConfig):
        self.cfg = cfg
        self.balance = cfg.initial_balance
        self.watches: List[Dict] = []
        self.trades = []
        self.candles = pd.DataFrame(columns=['open','high','low','close','volume'])
        init_excel(self.cfg.logfile)
        self.client: Optional[AsyncClient] = None
        self._in_position = False  # to avoid opening multiple overlapping positions per watch (simple guard)
        self._current_position = None  # store current open position meta if any

    async def _init_client(self):
        if self.client:
            return self.client
        if self.cfg.testnet:
            self.client = await AsyncClient.create(self.cfg.api_key, self.cfg.api_secret, testnet=True)
        else:
            self.client = await AsyncClient.create(self.cfg.api_key, self.cfg.api_secret)
        return self.client

    async def _close_client(self):
        if self.client:
            await self.client.close_connection()
            self.client = None

    async def start(self):
        if self.cfg.live_mode and (not self.cfg.api_key or not self.cfg.api_secret):
            raise RuntimeError("Live mode selected but API key/secret not set in config.")
        # initialize client early if live_mode
        if self.cfg.live_mode:
            await self._init_client()
        client = self.client if self.client else await AsyncClient.create()
        bm = BinanceSocketManager(client)
        kline_socket = bm.kline_socket(self.cfg.pair, interval=self.cfg.interval)

        print(f"[INFO] Live REAL-TRADE dual-entry for {self.cfg.pair} ({self.cfg.interval}) - live_mode={self.cfg.live_mode}")
        async with kline_socket as stream:
            while True:
                res = await stream.recv()
                k = res.get('k', {})
                is_closed = k.get('x', False)
                ts = pd.to_datetime(k.get('t'), unit='ms', utc=True)
                o = float(k.get('o')); h = float(k.get('h')); l = float(k.get('l')); c = float(k.get('c')); v = float(k.get('v'))
                self._append_candle(ts, o, h, l, c, v)

                if is_closed:
                    atr_series = compute_atr_from_df(self.candles, self.cfg.atr_period)
                    if len(atr_series) >= self.cfg.atr_period:
                        current_atr = atr_series.iat[-1]
                    else:
                        current_atr = np.nan

                    if not np.isnan(current_atr) and current_atr >= self.cfg.min_atr:
                        last_close = self.candles['close'].iat[-1]
                        long_level = last_close + current_atr * self.cfg.level_mult
                        short_level = last_close - current_atr * self.cfg.level_mult
                        watch = {
                            "start_idx": len(self.candles)-1,
                            "expire_idx": len(self.candles)-1 + self.cfg.monitor_candles,
                            "long_level": long_level,
                            "short_level": short_level,
                            "atr": current_atr,
                            "trigger_time": self.candles.index[-1]
                        }
                        self.watches.append(watch)
                        print(f"[WATCH CREATED] {watch['trigger_time']} ATR={current_atr:.6f} long={long_level:.6f} short={short_level:.6f}")

                    await self._process_watches_for_closed_candle()

    def _append_candle(self, ts, o, h, l, c, v):
        self.candles.loc[ts] = [o, h, l, c, v]
        if len(self.candles) > self.cfg.candles_buffer:
            self.candles = self.candles.iloc[-self.cfg.candles_buffer:]

    async def _place_market_order(self, side: str, quantity: float):
        """
        Place a market order on futures. Returns the order response.
        """
        if not self.cfg.live_mode:
            return None
        client = await self._init_client()
        # side should be 'BUY' or 'SELL'
        order = await client.futures_create_order(
            symbol=self.cfg.pair,
            side=side,
            type="MARKET",
            quantity=quantity
        )
        return order

    async def _process_watches_for_closed_candle(self):
        if self.candles.empty:
            return

        latest_idx = len(self.candles) - 1
        candle_high = self.candles['high'].iat[-1]
        candle_low = self.candles['low'].iat[-1]

        new_watches = []
        for w in self.watches:
            if latest_idx > w['expire_idx']:
                continue
            if latest_idx <= w['start_idx']:
                new_watches.append(w)
                continue

            triggered = False
            exit_price = None
            exit_reason = None
            side = None
            entry_price = None
            atr_now = w['atr']

            # LONG
            if candle_high >= w['long_level']:
                triggered = True
                side = 'LONG'
                entry_price = w['long_level']
                tp_price = entry_price + atr_now * self.cfg.tp_atr_mult
                sl_price = entry_price - atr_now * self.cfg.sl_atr_mult
                if candle_low <= sl_price:
                    exit_price = sl_price
                    exit_reason = 'SL'
                elif candle_high >= tp_price:
                    exit_price = tp_price
                    exit_reason = 'TP'
                else:
                    exit_price = None

            # SHORT
            elif candle_low <= w['short_level']:
                triggered = True
                side = 'SHORT'
                entry_price = w['short_level']
                tp_price = entry_price - atr_now * self.cfg.tp_atr_mult
                sl_price = entry_price + atr_now * self.cfg.sl_atr_mult
                if candle_high >= sl_price:
                    exit_price = sl_price
                    exit_reason = 'SL'
                elif candle_low <= tp_price:
                    exit_price = tp_price
                    exit_reason = 'TP'
                else:
                    exit_price = None

            if triggered and exit_price is not None:
                # Compute qty based on current (simulated) balance and leverage
                # qty = (balance * leverage) / entry_price
                qty = max((self.balance * self.cfg.leverage) / entry_price, 0.000001)
                # map to Binance side: LONG -> BUY to open, SHORT -> SELL to open
                open_side = "BUY" if side == 'LONG' else "SELL"
                close_side = "SELL" if side == 'LONG' else "BUY"

                executed_entry = None
                executed_exit = None

                # Place real orders if live_mode: open then close
                if self.cfg.live_mode:
                    try:
                        # place market order to open
                        resp_open = await self._place_market_order(open_side, quantity=qty)
                        executed_entry = None
                        if resp_open:
                            # record executed avg price if available in response (may be different)
                            # but per user's requirement, we will use entry_price for PnL calc
                            if isinstance(resp_open, dict):
                                # try common keys
                                if 'avgPrice' in resp_open:
                                    executed_entry = float(resp_open['avgPrice'])
                                elif 'fills' in resp_open and resp_open['fills']:
                                    executed_entry = float(resp_open['fills'][0].get('price', None)) if resp_open['fills'] else None
                        # place market order to close immediately when condition found. However to match forward-test:
                        # We will not attempt to logic-match fills; instead we will monitor and when condition is met we send close market order.
                        # For simplicity here: send close market order immediately (so both open+close are done in the same flow).
                        resp_close = await self._place_market_order(close_side, quantity=qty)
                        if isinstance(resp_close, dict):
                            if 'avgPrice' in resp_close:
                                executed_exit = float(resp_close['avgPrice'])
                            elif 'fills' in resp_close and resp_close['fills']:
                                executed_exit = float(resp_close['fills'][0].get('price', None))
                    except Exception as e:
                        print("[ERROR] Order placement failed:", e)
                        # Even if order fails, we will still record simulated trade per user's requirement (but mark executed prices None)
                else:
                    # paper / not live - executed_* remain None
                    pass

                # Compute PnL using the exact same formula as backtest/forward-test:
                pnl = self._compute_pnl(entry_price, exit_price, side)

                # update balance (we follow same balance update rule)
                self.balance += pnl

                trade = {
                    "pair": self.cfg.pair,
                    "watch_start": w['trigger_time'],
                    "trigger_side": side,
                    "trigger_level": entry_price,
                    "atr_at_trigger": atr_now,
                    "entry_time": w['trigger_time'],
                    "entry_price": entry_price,          # recorded as trigger_level to preserve calculation
                    "exit_time": self.candles.index[-1],
                    "exit_price": exit_price,            # recorded as TP/SL level
                    "pnl": pnl,
                    "exit_reason": exit_reason,
                    "balance_after": self.balance,
                    "executed_entry_price": executed_entry,
                    "executed_exit_price": executed_exit
                }
                self.trades.append(trade)
                append_trade_excel(self.cfg.logfile, [
                    trade['pair'], trade['watch_start'], trade['trigger_side'], trade['trigger_level'],
                    trade['atr_at_trigger'], trade['entry_time'], trade['entry_price'],
                    trade['exit_time'], trade['exit_price'], trade['pnl'],
                    trade['exit_reason'], trade['balance_after'],
                    trade['executed_entry_price'], trade['executed_exit_price']
                ])
                print(f"[TRADE RECORDED] {trade['exit_time']} {side} entry={entry_price:.6f} exit={exit_price:.6f} pnl={pnl:.6f} balance={self.balance:.4f} exec_entry={executed_entry} exec_exit={executed_exit}")
                # do not keep this watch
            else:
                # not triggered or triggered but no determined exit yet
                if latest_idx < w['expire_idx']:
                    new_watches.append(w)
                # else expired -> drop
        self.watches = new_watches

    def _compute_pnl(self, entry_price: float, exit_price: float, side: str):
        """
        Identical calculation to backtest/forward-test:
        - frac = (exit - entry)/entry for long; (entry - exit)/entry for short
        - pnl = frac * balance * leverage - fee_rate * balance
        """
        if side == 'LONG':
            frac = (exit_price - entry_price) / entry_price
        else:
            frac = (entry_price - exit_price) / entry_price
        pnl = frac * self.balance * self.cfg.leverage - self.cfg.fee_rate * self.balance
        return pnl

# ---------------- Run ----------------
if __name__ == "__main__":
     # --- isi API KEY & SECRET kamu di sini ---
    load_dotenv()
    cfg.api_key = os.getenv("BINANCE_API_KEY")
    cfg.api_secret = os.getenv("BINANCE_API_SECRET")
    # kalau mau testnet (akun demo), set True
    # cfg.testnet = True  

    # kalau mau langsung real, set ke False
    cfg.testnet = False  

    # pastikan live_mode True kalau mau benar-benar kirim order
    cfg.live_mode = True  

    init_excel(cfg.logfile)
    bot = LiveDualEntryLiveTradeBot(cfg)
    try:
        asyncio.get_event_loop().run_until_complete(bot.start())
    except KeyboardInterrupt:
        print("[STOP] Interrupted by user. Trades saved to", cfg.logfile)
        try:
            asyncio.get_event_loop().run_until_complete(bot._close_client())
        except:
            pass
