# live_bot_dual_entry_live.py
"""
Live trading (Binance USDT-M Futures)
- ATR dual-entry breakout (same rules as backtest)
- Single concurrent position per pair (self._current_position)
- Minimal holding time (min_hold_sec) before TP/SL applied
- Live_mode (True) will place real market orders; otherwise runs paper-mode
- Supports testnet via cfg.testnet
"""

import asyncio
import os
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import List, Dict, Optional

import numpy as np
import pandas as pd
from binance import AsyncClient, BinanceSocketManager
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()  # loads BINANCE_API_KEY, BINANCE_API_SECRET if present in .env

# ---------------- Config ----------------
@dataclass
class BotConfig:
    pair: str = "AVAXUSDT"
    interval: str = "1m"
    initial_balance: float = 20.0
    leverage: float = 3.0
    fee_rate: float = 0.0004
    min_atr: float = 0.0005
    atr_period: int = 14
    level_mult: float = 0.2
    tp_atr_mult: float = 0.8
    sl_atr_mult: float = 0.8
    monitor_candles: int = 3
    candles_buffer: int = 500
    min_hold_sec: int = 30
    logfile: str = "trades_log_live.xlsx"
    live_mode: bool = False        # True -> place real orders
    api_key: str = os.getenv("BINANCE_API_KEY")
    api_secret: str = os.getenv("BINANCE_API_SECRET")
    testnet: bool = False          # True -> AsyncClient.create(..., testnet=True)

cfg = BotConfig()

# ---------------- Excel helpers ----------------
def init_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append([
            "pair", "watch_start", "trigger_side", "trigger_level", "atr_at_trigger",
            "entry_time", "entry_price", "exit_time", "exit_price", "pnl",
            "exit_reason", "balance_after", "executed_entry_price", "executed_exit_price", "qty"
        ])
        wb.save(path)

def append_trade_excel(path: str, row: List):
    wb = load_workbook(path)
    ws = wb["Trades"]
    clean_row = []
    for v in row:
        if isinstance(v, datetime):
            clean_row.append(v.replace(tzinfo=None))
        else:
            clean_row.append(v)
    ws.append(clean_row)
    wb.save(path)

# ---------------- Technical helpers ----------------
def compute_atr_from_df(df: pd.DataFrame, period: int):
    high_low = df['high'] - df['low']
    high_close = (df['high'] - df['close'].shift(1)).abs()
    low_close = (df['low'] - df['close'].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    return atr

# ---------------- Live Bot ----------------
class LiveDualEntryLiveTrade:
    def __init__(self, cfg: BotConfig):
        self.cfg = cfg
        self.balance = cfg.initial_balance
        self.watches: List[Dict] = []
        self.candles = pd.DataFrame(columns=['open','high','low','close','volume'])
        self._current_position: Optional[Dict] = None
        self.client: Optional[AsyncClient] = None
        init_excel(self.cfg.logfile)

    async def _init_client(self):
        if self.client:
            return self.client
        # create async client (testnet param supported)
        if self.cfg.testnet:
            self.client = await AsyncClient.create(self.cfg.api_key, self.cfg.api_secret, testnet=True)
        else:
            self.client = await AsyncClient.create(self.cfg.api_key, self.cfg.api_secret)
        # set leverage on the symbol if live_mode
        try:
            await self.client.futures_change_leverage(symbol=self.cfg.pair, leverage=self.cfg.leverage)
            print(f"[INFO] Leverage set to {self.cfg.leverage}")
        except Exception as e:
            print("[WARN] set leverage failed:", e)
        return self.client

    async def _close_client(self):
        if self.client:
            await self.client.close_connection()
            self.client = None

    async def _format_quantity(self, qty: float) -> float:
        """Format qty to symbol lot size; returns rounded qty (float)."""
        client = await self._init_client()
        try:
            info = await client.futures_exchange_info()
            for s in info['symbols']:
                if s['symbol'] == self.cfg.pair:
                    for f in s['filters']:
                        if f['filterType'] == 'LOT_SIZE':
                            step_size = float(f['stepSize'])
                            min_qty = float(f['minQty'])
                            # precision from step_size
                            precision = int(round(-np.log10(step_size)))
                            qty = max(min_qty, round(qty, precision))
                            return qty
        except Exception as e:
            print("[WARN] format qty failed:", e)
        return qty

    async def _place_market_order(self, side: str, qty: float):
        """Place a market order on futures (returns order response dict)"""
        if not self.cfg.live_mode:
            return None
        client = await self._init_client()
        try:
            resp = await client.futures_create_order(
                symbol=self.cfg.pair,
                side=side,
                type="MARKET",
                quantity=qty
            )
            return resp
        except Exception as e:
            print("[ERROR] placing market order failed:", e)
            return None

    # main start
    async def start(self):
        if self.cfg.live_mode and (not self.cfg.api_key or not self.cfg.api_secret):
            raise RuntimeError("Live mode selected but API key/secret not set.")
        # init client only for socket (even if paper mode, AsyncClient used)
        await self._init_client()
        client = self.client
        bm = BinanceSocketManager(client)
        kline_socket = bm.kline_socket(self.cfg.pair, interval=self.cfg.interval)

        print(f"[INFO] Starting live trade bot for {self.cfg.pair} interval={self.cfg.interval} live_mode={self.cfg.live_mode} testnet={self.cfg.testnet}")
        async with kline_socket as stream:
            while True:
                res = await stream.recv()
                k = res.get('k', {})
                is_closed = k.get('x', False)
                ts = pd.to_datetime(k.get('t'), unit='ms', utc=True)
                o = float(k.get('o')); h = float(k.get('h')); l = float(k.get('l')); c = float(k.get('c')); v = float(k.get('v'))
                self._append_candle(ts, o, h, l, c, v)

                if is_closed:
                    atr_series = compute_atr_from_df(self.candles, self.cfg.atr_period)
                    current_atr = atr_series.iat[-1] if len(atr_series) >= self.cfg.atr_period else np.nan

                    if not np.isnan(current_atr) and current_atr >= self.cfg.min_atr:
                        # create watch only if currently no open position (guard)
                        if self._current_position is None:
                            last_close = self.candles['close'].iat[-1]
                            long_level = last_close + current_atr * self.cfg.level_mult
                            short_level = last_close - current_atr * self.cfg.level_mult
                            watch = {
                                "start_idx": len(self.candles)-1,
                                "expire_idx": len(self.candles)-1 + self.cfg.monitor_candles,
                                "long_level": long_level,
                                "short_level": short_level,
                                "atr": current_atr,
                                "trigger_time": self.candles.index[-1]
                            }
                            self.watches.append(watch)
                            print(f"[WATCH CREATED] {watch['trigger_time']} ATR={current_atr:.6f} long={long_level:.6f} short={short_level:.6f}")

                    # process watches -> may open a live position (market order) if triggered
                    await self._process_watches_and_maybe_open()

                    # if there is a current position, check for exit conditions (after min_hold_sec)
                    await self._process_current_position_and_maybe_close()

    def _append_candle(self, ts, o, h, l, c, v):
        self.candles.loc[ts] = [o, h, l, c, v]
        if len(self.candles) > self.cfg.candles_buffer:
            self.candles = self.candles.iloc[-self.cfg.candles_buffer:]

    async def _process_watches_and_maybe_open(self):
        if self._current_position is not None:
            return  # guard: one position at a time

        latest_idx = len(self.candles) - 1
        candle_high = self.candles['high'].iat[-1]
        candle_low = self.candles['low'].iat[-1]

        new_watches = []
        for w in self.watches:
            if latest_idx <= w['start_idx']:
                new_watches.append(w)
                continue

            triggered = False
            side = None
            entry_price = None
            atr_now = w['atr']

            if candle_high >= w['long_level']:
                triggered = True
                side = 'LONG'
                entry_price = w['long_level']
            elif candle_low <= w['short_level']:
                triggered = True
                side = 'SHORT'
                entry_price = w['short_level']

            if triggered:
                tp_price = entry_price + atr_now * self.cfg.tp_atr_mult if side == 'LONG' else entry_price - atr_now * self.cfg.tp_atr_mult
                sl_price = entry_price - atr_now * self.cfg.sl_atr_mult if side == 'LONG' else entry_price + atr_now * self.cfg.sl_atr_mult

                # compute quantity: (balance * leverage) / entry_price (approx USD exposure)
                qty = max((self.balance * self.cfg.leverage) / entry_price, 0.000001)
                qty = await self._format_quantity(qty)

                executed_entry_price = None
                executed_exit_price = None

                # open market order if live_mode
                if self.cfg.live_mode:
                    open_side = "BUY" if side == 'LONG' else "SELL"
                    resp_open = await self._place_market_order(open_side, qty)
                    if resp_open and isinstance(resp_open, dict):
                        # try extract executed price from resp
                        executed_entry_price = float(resp_open.get('avgPrice') or (resp_open.get('fills',[{}])[0].get('price', None) or 0))

                # set current position metadata (use trigger-level as entry_price for PnL calc to match backtest)
                self._current_position = {
                    "side": side,
                    "entry_price": entry_price,
                    "executed_entry_price": executed_entry_price,
                    "tp_price": tp_price,
                    "sl_price": sl_price,
                    "qty": qty,
                    "entry_time": w['trigger_time'],
                    "atr": atr_now
                }
                print(f"[OPENED] {side} entry_level={entry_price:.6f} tp={tp_price:.6f} sl={sl_price:.6f} qty={qty} exec_entry={executed_entry_price}")
                # do not add watch back
            else:
                if latest_idx < w['expire_idx']:
                    new_watches.append(w)
        self.watches = new_watches

    async def _process_current_position_and_maybe_close(self):
        if self._current_position is None:
            return

        pos = self._current_position
        latest_candle = self.candles.iloc[-1]
        now = self.candles.index[-1]
        elapsed = (now - pos['entry_time']).total_seconds()

        exit_reason = None
        exit_price = None
        executed_exit_price = None

        # only consider TP/SL after min_hold_sec
        if elapsed >= self.cfg.min_hold_sec:
            if pos['side'] == 'LONG':
                if latest_candle['high'] >= pos['tp_price']:
                    exit_price = pos['tp_price']; exit_reason = 'TP'
                elif latest_candle['low'] <= pos['sl_price']:
                    exit_price = pos['sl_price']; exit_reason = 'SL'
            else:  # SHORT
                if latest_candle['low'] <= pos['tp_price']:
                    exit_price = pos['tp_price']; exit_reason = 'TP'
                elif latest_candle['high'] >= pos['sl_price']:
                    exit_price = pos['sl_price']; exit_reason = 'SL'

        if exit_price is not None:
            # place close market order if live_mode
            if self.cfg.live_mode:
                close_side = "SELL" if pos['side']=='LONG' else "BUY"
                resp_close = await self._place_market_order(close_side, pos['qty'])
                if resp_close and isinstance(resp_close, dict):
                    executed_exit_price = float(resp_close.get('avgPrice') or (resp_close.get('fills',[{}])[0].get('price', None) or 0))

            # compute pnl using trigger-level entry_price to match backtest
            pnl = self._compute_pnl(pos['entry_price'], exit_price, pos['side'])
            self.balance += pnl

            trade = [
                self.cfg.pair, pos['entry_time'], pos['side'], pos['entry_price'], pos['atr'],
                pos['entry_time'], pos['entry_price'], now, exit_price, pnl,
                exit_reason, self.balance, pos.get('executed_entry_price'), executed_exit_price, pos['qty']
            ]
            append_trade_excel(self.cfg.logfile, trade)
            print(f"[CLOSED] {now} {pos['side']} entry={pos['entry_price']:.6f} exit={exit_price:.6f} pnl={pnl:.6f} balance={self.balance:.4f} exec_exit={executed_exit_price}")
            self._current_position = None

    def _compute_pnl(self, entry_price: float, exit_price: float, side: str):
        if side == 'LONG':
            frac = (exit_price - entry_price) / entry_price
        else:
            frac = (entry_price - exit_price) / entry_price
        pnl = frac * self.balance * self.cfg.leverage - self.cfg.fee_rate * self.balance
        return pnl

# ---------------- Run ----------------
if __name__ == "__main__":
    # config quick toggles:
    cfg.testnet = False      # set True to use Binance futures testnet (recommended initial)
    cfg.live_mode = False   # set True to place real orders (only after verified)
    # API keys read from .env by default (BINANCE_API_KEY, BINANCE_API_SECRET)
    # Or set directly:
    # cfg.api_key = "..." 
    # cfg.api_secret = "..."

    load_dotenv()
    cfg.api_key = os.getenv("BINANCE_API_KEY")
    cfg.api_secret = os.getenv("BINANCE_API_SECRET")

    init_excel(cfg.logfile)
    bot = LiveDualEntryLiveTrade(cfg)
    try:
        asyncio.get_event_loop().run_until_complete(bot.start())
    except KeyboardInterrupt:
        print("[STOP] Interrupted by user. Trades saved to", cfg.logfile)
        try:
            asyncio.get_event_loop().run_until_complete(bot._close_client())
        except:
            pass
