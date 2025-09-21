# live_bot_dual_entry_v2_realistic.py
import asyncio
import os
import random
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Optional, Tuple

import numpy as np
import pandas as pd
from binance import AsyncClient, BinanceSocketManager
from openpyxl import Workbook, load_workbook
from binance.exceptions import BinanceWebsocketException
import aiohttp

# ---------------- Config ----------------
@dataclass
class BotConfig:
    pair: str = "AVAXUSDT"
    interval: str = "1m"
    initial_balance: float = 20.0
    leverage: float = 5.0  # Reduced from 20.0 to 5.0 for more realistic risk
    fee_rate: float = 0.0004  # 0.04% fee
    min_atr: float = 0.0005
    atr_period: int = 14
    level_mult: float = 0.2
    tp_atr_mult: float = 1.5  # Increased from 0.9 for better risk/reward
    sl_atr_mult: float = 1.0  # Adjusted from 0.9
    monitor_candles: int = 3
    candles_buffer: int = 500
    min_hold_sec: int = 30    # minimal holding time sebelum exit
    logfile: str = "trades_log.xlsx"
    
    # New realistic parameters
    slippage_factor: float = 0.0003  # 0.03% slippage
    execution_probability: float = 0.85  # 85% chance of execution at desired price
    min_volume_ratio: float = 1.2  # Minimum volume ratio vs 20-period average
    max_position_pct: float = 0.1  # Maximum 10% of balance per trade

cfg = BotConfig()

# ---------------- Excel helpers ----------------
def init_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append([
            "pair", "watch_start", "trigger_side", "trigger_level", "atr_at_trigger",
            "entry_time", "entry_price", "exit_time", "exit_price", "pnl",
            "exit_reason", "balance_after", "slippage_entry", "slippage_exit"
        ])
        wb.save(path)

def append_trade_excel(path: str, row: List):
    wb = load_workbook(path)
    ws = wb["Trades"]
    clean_row = [v.replace(tzinfo=None) if isinstance(v, datetime) else v for v in row]
    ws.append(clean_row)
    wb.save(path)

# ---------------- Technical helpers ----------------
def compute_atr_from_df(df: pd.DataFrame, period: int):
    high_low = df['high'] - df['low']
    high_close = (df['high'] - df['close'].shift(1)).abs()
    low_close = (df['low'] - df['close'].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    return atr

def calculate_volume_ratio(df: pd.DataFrame, period: int = 20):
    """Calculate current volume vs period average"""
    if len(df) < period:
        return 1.0
    current_volume = df['volume'].iloc[-1]
    avg_volume = df['volume'].rolling(period).mean().iloc[-1]
    return current_volume / avg_volume if avg_volume > 0 else 1.0

def calculate_volatility(df: pd.DataFrame, period: int = 20):
    """Calculate recent volatility as standard deviation of returns"""
    if len(df) < period:
        return 0.0
    returns = df['close'].pct_change()
    return returns.rolling(period).std().iloc[-1]

# ---------------- Realistic execution simulation ----------------
def simulate_execution(desired_price: float, side: str, current_candle: pd.Series, 
                      slippage_factor: float, execution_prob: float) -> Tuple[Optional[float], float]:
    """
    Simulate realistic order execution with slippage and execution probability
    
    Returns:
        Tuple of (executed_price, slippage_amount) or (None, 0) if not executed
    """
    # Check if we get executed based on probability
    if random.random() > execution_prob:
        return None, 0
    
    # Calculate slippage based on side and market conditions
    slippage_pct = slippage_factor * (1 + random.random())  # Random additional slippage
    
    if side == 'LONG':
        # For long orders, we pay more than desired price
        executed_price = desired_price * (1 + slippage_pct)
        # Check if price actually reached this level (high must be >= executed_price)
        if current_candle['high'] < executed_price:
            return None, 0
    else:  # SHORT
        # For short orders, we get less than desired price
        executed_price = desired_price * (1 - slippage_pct)
        # Check if price actually reached this level (low must be <= executed_price)
        if current_candle['low'] > executed_price:
            return None, 0
    
    slippage_amount = abs(executed_price - desired_price)
    return executed_price, slippage_amount

# ---------------- Live Bot ----------------
class LiveDualEntryBot:
    def __init__(self, cfg: BotConfig):
        self.cfg = cfg
        self.balance = cfg.initial_balance
        self.candles = pd.DataFrame(columns=['open','high','low','close','volume'])
        self._current_position: Optional[Dict] = None
        self.watches: List[Dict] = []
        init_excel(self.cfg.logfile)

    async def start(self):
        max_retries = 5
        retry_delay = 5  # seconds
        
        for attempt in range(max_retries):
            try:
                client = await AsyncClient.create()
                bm = BinanceSocketManager(client)
                
                # Set timeout untuk mencegah hanging connection
                async with bm.kline_socket(self.cfg.pair, interval=self.cfg.interval) as stream:
                    print(f"[INFO] Connected to WebSocket (Attempt {attempt + 1}/{max_retries})")
                    print(f"[INFO] Live dual-entry forward-test for {self.cfg.pair}")
                    
                    while True:
                        try:
                            # Tambahkan timeout untuk receive data
                            res = await asyncio.wait_for(stream.recv(), timeout=30.0)
                            k = res.get('k', {})
                            
                            # Pastikan data tidak None sebelum processing
                            if not k:
                                continue
                                
                            is_closed = k.get('x', False)
                            ts = k.get('t')
                            o, h, l, c, v = k.get('o'), k.get('h'), k.get('l'), k.get('c'), k.get('v')
                            
                            # Validasi data sebelum processing
                            if None in [ts, o, h, l, c, v]:
                                print("[WARNING] Received incomplete candle data, skipping...")
                                continue
                                
                            ts = pd.to_datetime(ts, unit='ms', utc=True)
                            o, h, l, c, v = map(float, (o, h, l, c, v))
                            
                            self._append_candle(ts, o, h, l, c, v)
                            
                            if is_closed:
                                print(f"Candle closed at {ts} - Processing strategies...")
                                await self._process_strategies()
                                
                        except asyncio.TimeoutError:
                            print("[WARNING] WebSocket timeout, reconnecting...")
                            break
                        except Exception as e:
                            print(f"[ERROR] Processing error: {e}")
                            continue
                            
            except (BinanceWebsocketException, aiohttp.ClientError, ConnectionError) as e:
                print(f"[ERROR] Connection failed (Attempt {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    print(f"Retrying in {retry_delay} seconds...")
                    await asyncio.sleep(retry_delay)
                    retry_delay *= 2  # Exponential backoff
                else:
                    print("[CRITICAL] Max retries exceeded. Exiting...")
                    break
            except Exception as e:
                print(f"[CRITICAL] Unexpected error: {e}")
                break

    def _append_candle(self, ts, o, h, l, c, v):
        self.candles.loc[ts] = [o, h, l, c, v]
        if len(self.candles) > self.cfg.candles_buffer:
            self.candles = self.candles.iloc[-self.cfg.candles_buffer:]

    def _create_watch(self, atr_value, volume_ratio, volatility):
        if self._current_position is not None:
            return
            
        last_close = self.candles['close'].iat[-1]
        watch = {
            "start_idx": len(self.candles)-1,
            "expire_idx": len(self.candles)-1 + self.cfg.monitor_candles,
            "long_level": last_close + atr_value * self.cfg.level_mult,
            "short_level": last_close - atr_value * self.cfg.level_mult,
            "atr": atr_value,
            "trigger_time": self.candles.index[-1],
            "volume_ratio": volume_ratio,
            "volatility": volatility
        }
        self.watches.append(watch)
        print(f"[WATCH CREATED] {watch['trigger_time']} ATR={atr_value:.6f} VolRatio={volume_ratio:.2f} Volatility={volatility:.6f}")

    def _process_watches(self):
        if self._current_position is not None:
            return

        latest_idx = len(self.candles) - 1
        current_candle = self.candles.iloc[-1]

        new_watches = []
        for w in self.watches:
            if latest_idx <= w['start_idx']:
                new_watches.append(w)
                continue

            triggered = False
            side, entry_price, slippage_entry = None, None, 0
            tp_price, sl_price = None, None

            # Check long trigger with realistic execution
            if current_candle['high'] >= w['long_level']:
                entry_price, slippage_entry = simulate_execution(
                    w['long_level'], 'LONG', current_candle, 
                    self.cfg.slippage_factor, self.cfg.execution_probability
                )
                if entry_price is not None:
                    triggered = True
                    side = 'LONG'
                    tp_price = entry_price + w['atr'] * self.cfg.tp_atr_mult
                    sl_price = entry_price - w['atr'] * self.cfg.sl_atr_mult

            # Check short trigger with realistic execution
            elif current_candle['low'] <= w['short_level']:
                entry_price, slippage_entry = simulate_execution(
                    w['short_level'], 'SHORT', current_candle, 
                    self.cfg.slippage_factor, self.cfg.execution_probability
                )
                if entry_price is not None:
                    triggered = True
                    side = 'SHORT'
                    tp_price = entry_price - w['atr'] * self.cfg.tp_atr_mult
                    sl_price = entry_price + w['atr'] * self.cfg.sl_atr_mult

            if triggered:
                # Calculate position size with risk management
                position_size = min(self.balance * self.cfg.max_position_pct, 
                                   self.balance * self.cfg.leverage)
                
                self._current_position = {
                    "side": side,
                    "entry_price": entry_price,
                    "tp_price": tp_price,
                    "sl_price": sl_price,
                    "entry_time": w['trigger_time'],
                    "atr": w['atr'],
                    "slippage_entry": slippage_entry,
                    "position_size": position_size,
                    "volume_ratio": w['volume_ratio'],
                    "volatility": w['volatility']
                }
                print(f"[POSITION OPENED] {side} at {entry_price:.6f}, TP: {tp_price:.6f}, SL: {sl_price:.6f}")
            else:
                if latest_idx < w['expire_idx']:
                    new_watches.append(w)
                    
        self.watches = new_watches

    def _process_current_position(self):
        if self._current_position is None:
            return

        pos = self._current_position
        current_candle = self.candles.iloc[-1]
        now = self.candles.index[-1]
        elapsed_sec = (now - pos['entry_time']).total_seconds()

        exit_reason = None
        exit_price = None
        slippage_exit = 0

        # TP/SL hanya diperiksa jika sudah lewat minimal holding time
        if elapsed_sec >= self.cfg.min_hold_sec:
            if pos['side'] == 'LONG':
                # Check TP
                if current_candle['high'] >= pos['tp_price']:
                    exit_price, slippage_exit = simulate_execution(
                        pos['tp_price'], 'LONG', current_candle,
                        self.cfg.slippage_factor, self.cfg.execution_probability
                    )
                    if exit_price is not None:
                        exit_reason = 'TP'
                # Check SL
                elif current_candle['low'] <= pos['sl_price'] and exit_price is None:
                    exit_price, slippage_exit = simulate_execution(
                        pos['sl_price'], 'LONG', current_candle,
                        self.cfg.slippage_factor, self.cfg.execution_probability
                    )
                    if exit_price is not None:
                        exit_reason = 'SL'
            else:  # SHORT
                # Check TP
                if current_candle['low'] <= pos['tp_price']:
                    exit_price, slippage_exit = simulate_execution(
                        pos['tp_price'], 'SHORT', current_candle,
                        self.cfg.slippage_factor, self.cfg.execution_probability
                    )
                    if exit_price is not None:
                        exit_reason = 'TP'
                # Check SL
                elif current_candle['high'] >= pos['sl_price'] and exit_price is None:
                    exit_price, slippage_exit = simulate_execution(
                        pos['sl_price'], 'SHORT', current_candle,
                        self.cfg.slippage_factor, self.cfg.execution_probability
                    )
                    if exit_price is not None:
                        exit_reason = 'SL'

        if exit_price is not None:
            pnl = self._compute_pnl(pos['entry_price'], exit_price, pos['side'], pos['position_size'])
            self.balance += pnl
            trade = {
                "pair": self.cfg.pair,
                "watch_start": pos['entry_time'],
                "trigger_side": pos['side'],
                "trigger_level": pos['entry_price'] - (pos['slippage_entry'] if pos['side'] == 'LONG' else -pos['slippage_entry']),
                "atr_at_trigger": pos['atr'],
                "entry_time": pos['entry_time'],
                "entry_price": pos['entry_price'],
                "exit_time": now,
                "exit_price": exit_price,
                "pnl": pnl,
                "exit_reason": exit_reason,
                "balance_after": self.balance,
                "slippage_entry": pos['slippage_entry'],
                "slippage_exit": slippage_exit
            }
            append_trade_excel(self.cfg.logfile, [
                trade['pair'], trade['watch_start'], trade['trigger_side'], trade['trigger_level'],
                trade['atr_at_trigger'], trade['entry_time'], trade['entry_price'],
                trade['exit_time'], trade['exit_price'], trade['pnl'],
                trade['exit_reason'], trade['balance_after'],
                trade['slippage_entry'], trade['slippage_exit']
            ])
            print(f"[TRADE CLOSED] {now} {pos['side']} Entry={pos['entry_price']:.6f} Exit={exit_price:.6f} "
                  f"PnL={pnl:.6f} Balance={self.balance:.4f} Reason={exit_reason}")
            self._current_position = None

    def _compute_pnl(self, entry_price: float, exit_price: float, side: str, position_size: float):
        if side == 'LONG':
            price_change_pct = (exit_price - entry_price) / entry_price
        else:
            price_change_pct = (entry_price - exit_price) / entry_price
        
        # Calculate gross PnL
        gross_pnl = price_change_pct * position_size
        
        # Calculate fees (both entry and exit)
        trade_value = position_size / self.cfg.leverage  # Actual capital used
        fees = trade_value * self.cfg.fee_rate * 2  # Entry and exit fees
        
        # Net PnL
        net_pnl = gross_pnl - fees
        return net_pnl

# ---------------- Run ----------------
if __name__ == "__main__":
    init_excel(cfg.logfile)
    bot = LiveDualEntryBot(cfg)
    try:
        asyncio.get_event_loop().run_until_complete(bot.start())
    except KeyboardInterrupt:
        print("[STOP] Interrupted by user. Trades saved to", cfg.logfile)