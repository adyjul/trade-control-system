# live_trade_bot7_tick_enhanced.py
import asyncio
import math
import os
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Optional, Deque
from collections import deque
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from binance import AsyncClient, BinanceSocketManager
from binance.enums import SIDE_BUY, SIDE_SELL, ORDER_TYPE_MARKET, ORDER_TYPE_LIMIT, TIME_IN_FORCE_GTC
from dotenv import load_dotenv
import talib

load_dotenv()

# ---------------- Enhanced Config ----------------  
@dataclass
class BotConfig:
    # Existing config
    api_key: str = os.getenv('BINANCE_API_KEY')
    api_secret: str = os.getenv('BINANCE_API_SECRET')
    pair: str = "AVAXUSDT"
    interval: str = "5m"
    initial_balance: float = 20.0
    leverage: float = 10.0
    fee_rate: float = 0.0004
    min_atr: float = 0.0005
    atr_period: int = 14
    level_mult: float = 0.15
    tp_atr_mult: float = 0.9
    sl_atr_mult: float = 1.0
    monitor_candles: int = 3
    candles_buffer: int = 1000
    min_hold_sec: int = 600
    logfile: str = "trades_log.xlsx"
    risk_pct: float = 0.008
    margin_type: str = "ISOLATED"
    use_testnet: bool = True
    use_limit_orders: bool = True
    slippage_pct: float = 0.001
    require_confirmation: bool = True
    adaptive_tp_sl: bool = True

    # AVAX-specific precision settings
    qty_precision: int = 1
    price_precision: int = 3
    profit_pct_close_tp: float = 0.05

    # Guard settings
    max_hold_guard_sec = 600
    guard_profit_trigger = 0.04
    max_hold_guard_sec2 = 1200
    guard_profit_trigger2 = 0.03
    max_hold_guard_sec3 = 1800
    guard_profit_trigger3 = 0.02
    max_hold_guard_sl_sec = 1800
    guard_loss_trigger = -0.04

    daily_profit_lock_pct = 1.0
    daily_reset_hour = 7

    # NEW: Tick trading settings
    enable_tick_trading: bool = True  # Enable/disable tick trading
    tick_buffer_size: int = 50  # Store last 50 ticks for analysis
    min_tick_volume: float = 1000.0  # Minimum volume to consider tick valid
    tick_momentum_period: int = 10  # Period for tick momentum calculation
    quick_tp_pct: float = 0.005  # 0.5% quick TP for scalping
    quick_sl_pct: float = 0.003  # 0.3% quick SL for scalping
    max_scalp_duration: int = 300  # Max 5 minutes per scalp trade
    tick_volume_threshold: float = 2.0  # Volume multiplier for significant ticks

# ---------------- Excel Logger (Tetap Sama) ----------------
def init_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append([
            "pair", "entry_time", "side", "entry_price", "qty", "tp_price", "sl_price",
            "exit_time", "exit_price", "pnl", "fees", "exit_reason", "balance_after",
            "atr_value", "volatility_ratio", "trade_type"  # NEW: trade_type column
        ])
        wb.save(path)

def append_trade_excel(path: str, row: List):
    wb = load_workbook(path)
    ws = wb["Trades"]
    clean_row = [v.replace(tzinfo=None) if isinstance(v, datetime) else v for v in row]
    ws.append(clean_row)
    wb.save(path)

# ---------------- Technical Helpers (Tetap Sama) ----------------
def compute_atr_from_df(df: pd.DataFrame, period: int):
    high_low = df['high'] - df['low']
    high_close = (df['high'] - df['close'].shift(1)).abs()
    low_close = (df['low'] - df['close'].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    return atr

def adjust_qty(qty, step_size):
        # Membulatkan ke bawah sesuai step size
        return math.floor(qty / step_size) * step_size

def compute_volatility_ratio(df: pd.DataFrame, atr_period: int = 14):
    atr = compute_atr_from_df(df, atr_period)
    current_atr = atr.iat[-1] if len(atr) > 0 else 0
    price = df['close'].iat[-1] if len(df) > 0 else 1
    return current_atr / price if price > 0 else 0

def compute_qty_by_risk(balance: float, risk_pct: float, entry_price: float, sl_price: float):
    risk_amount = balance * risk_pct
    risk_per_unit = abs(entry_price - sl_price)
    if risk_per_unit == 0:
        return 0
    qty = risk_amount / risk_per_unit
    return qty

async def round_qty_to_step(client: AsyncClient, symbol: str, raw_qty: float):
    info = await client.futures_exchange_info()
    stepSize = None
    minQty = None
    for s in info['symbols']:
        if s['symbol'] == symbol:
            lot = next((f for f in s['filters'] if f['filterType'] == 'LOT_SIZE'), None)
            if lot:
                stepSize = float(lot['stepSize'])
                minQty = float(lot['minQty'])
            break
    if stepSize is None:
        raise RuntimeError("LOT_SIZE not found for symbol")
    rounded = math.floor(raw_qty / stepSize) * stepSize
    if rounded < (minQty or 0):
        return 0.0
    return float(f"{rounded:.8f}")

# ---------------- Enhanced Live Bot dengan Tick Trading ----------------
class ImprovedLiveDualEntryBot:
    def __init__(self, cfg: BotConfig):
        self.cfg = cfg
        self.balance = cfg.initial_balance
        self.candles = pd.DataFrame(columns=['open', 'high', 'low', 'close', 'volume'])
        
        # Existing attributes
        self._current_position: Optional[Dict] = None
        self.watches: List[Dict] = []
        self.pending_orders: List[Dict] = []
        self.active_orders: List[Dict] = []
        self._current_signal_side = None
        self.daily_start_equity = None
        self.daily_realized_pct = 0.0
        self.trade_locked = False
        self.volatility_ratio = 0.0
        self.order_timeout = 600
        
        # NEW: Tick trading attributes
        self.tick_prices: Deque[float] = deque(maxlen=cfg.tick_buffer_size)
        self.tick_volumes: Deque[float] = deque(maxlen=cfg.tick_buffer_size)
        self.tick_timestamps: Deque[datetime] = deque(maxlen=cfg.tick_buffer_size)
        self.last_tick_price = 0.0
        self.last_tick_time = None
        self.tick_analysis_ready = False
        
        init_excel(self.cfg.logfile)
        self.client: Optional[AsyncClient] = None
        self.bm: Optional[BinanceSocketManager] = None

    # NEW: Tick Analysis Methods
    def _calculate_tick_momentum(self) -> float:
        """Calculate momentum from recent ticks"""
        if len(self.tick_prices) < self.cfg.tick_momentum_period:
            return 0.0
        
        recent_prices = list(self.tick_prices)
        if len(recent_prices) < 2:
            return 0.0
            
        price_change = (recent_prices[-1] - recent_prices[0]) / recent_prices[0]
        return price_change

    def _calculate_tick_volatility(self) -> float:
        """Calculate volatility from recent ticks"""
        if len(self.tick_prices) < 10:
            return 0.0
        
        returns = []
        for i in range(1, len(self.tick_prices)):
            if self.tick_prices[i-1] > 0:
                ret = (self.tick_prices[i] - self.tick_prices[i-1]) / self.tick_prices[i-1]
                returns.append(abs(ret))
        
        return np.mean(returns) if returns else 0.0

    def _detect_tick_volume_spike(self) -> bool:
        """Detect significant volume spikes in tick data"""
        if len(self.tick_volumes) < 10:
            return False
        
        recent_volumes = list(self.tick_volumes)
        avg_volume = np.mean(recent_volumes[:-3])  # Average excluding last 3 ticks
        current_volume = recent_volumes[-1] if recent_volumes else 0
        
        return current_volume > avg_volume * self.cfg.tick_volume_threshold

    def _analyze_tick_sentiment(self) -> Dict[str, float]:
        """Analyze market sentiment from tick data"""
        sentiment = {
            "momentum": 0.0,
            "volatility": 0.0,
            "volume_spike": False,
            "trend_strength": 0.0
        }
        
        if len(self.tick_prices) >= 10:
            sentiment["momentum"] = self._calculate_tick_momentum()
            sentiment["volatility"] = self._calculate_tick_volatility()
            sentiment["volume_spike"] = self._detect_tick_volume_spike()
            
            # Simple trend strength (price consistency in direction)
            if len(self.tick_prices) >= 5:
                recent_changes = [self.tick_prices[i] - self.tick_prices[i-1] 
                                for i in range(1, len(self.tick_prices))]
                positive_changes = sum(1 for change in recent_changes if change > 0)
                sentiment["trend_strength"] = abs(positive_changes / len(recent_changes) - 0.5) * 2
        
        return sentiment

    # NEW: Tick-based Entry Conditions
    def _check_tick_entry_signal(self, current_price: float, is_buyer_maker: bool) -> Optional[str]:
        """Check for entry signals based on tick data"""
        if not self.tick_analysis_ready or self._current_position is not None:
            return None
        
        sentiment = self._analyze_tick_sentiment()
        
        # Strong bullish signal: positive momentum + high volume + low volatility
        bullish_condition = (
            sentiment["momentum"] > 0.001 and 
            sentiment["volume_spike"] and 
            sentiment["volatility"] < 0.002 and
            not is_buyer_maker  # Buyer is aggressive
        )
        
        # Strong bearish signal: negative momentum + high volume + low volatility  
        bearish_condition = (
            sentiment["momentum"] < -0.001 and
            sentiment["volume_spike"] and
            sentiment["volatility"] < 0.002 and
            is_buyer_maker  # Seller is aggressive
        )
        
        if bullish_condition:
            return "LONG"
        elif bearish_condition:
            return "SHORT"
        
        return None

    # NEW: Quick Scalping Methods
    async def _execute_quick_scalp(self, side: str, current_price: float):
        """Execute quick scalp trade based on tick signals"""
        if self._current_position is not None:
            return
            
        # Calculate smaller position size for scalping
        base_qty = self.calculate_proper_position_size(
            current_price, 
            current_price * (0.99 if side == "LONG" else 1.01)  # Conservative SL for calculation
        )
        scalp_qty = base_qty * 0.5  # 50% of normal size for scalping
        
        if scalp_qty <= 0:
            return
            
        # Set tight TP/SL for scalping
        if side == "LONG":
            tp_price = current_price * (1 + self.cfg.quick_tp_pct)
            sl_price = current_price * (1 - self.cfg.quick_sl_pct)
        else:
            tp_price = current_price * (1 - self.cfg.quick_tp_pct)
            sl_price = current_price * (1 + self.cfg.quick_sl_pct)
            
        # Round prices
        current_price = self._round_price(current_price)
        tp_price = self._round_price(tp_price)
        sl_price = self._round_price(sl_price)
        scalp_qty = self._round_qty(scalp_qty)
        
        try:
            # Use market order for fastest execution
            order = await self.client.futures_create_order(
                symbol=self.cfg.pair,
                side=SIDE_BUY if side == 'LONG' else SIDE_SELL,
                type=ORDER_TYPE_MARKET,
                quantity=scalp_qty
            )
            
            # Get execution price from order fills
            exec_price = current_price
            if 'fills' in order and order['fills']:
                filled_prices = [float(fill['price']) for fill in order['fills']]
                exec_price = sum(filled_prices) / len(filled_prices)
            
            exec_price = self._round_price(exec_price)
            
            # Update position
            self._current_position = {
                "side": side,
                "entry_price": exec_price,
                "qty": scalp_qty,
                "tp_price": tp_price,
                "sl_price": sl_price,
                "entry_time": datetime.now(timezone.utc),
                "atr": 0.0,  # Not used for scalping
                "vol_mult": 1.0,
                "is_scalp": True,  # Mark as scalp trade
                "scalp_start_time": datetime.now(timezone.utc)
            }
            
            print(f"🎯 [QUICK SCALP] {side} {scalp_qty:.2f} @ {exec_price:.3f} | TP: {tp_price:.3f} | SL: {sl_price:.3f}")
            
            # Log scalp trade
            append_trade_excel(self.cfg.logfile, [
                self.cfg.pair,
                datetime.now(timezone.utc),
                side,
                exec_price,
                scalp_qty,
                tp_price,
                sl_price,
                None,  # exit_time
                None,  # exit_price
                0,  # pnl
                0,  # fees
                "SCALP_ENTRY",
                self.balance,
                0.0,  # atr_value
                0.0,  # volatility_ratio
                "SCALP"  # trade_type
            ])
            
        except Exception as e:
            print(f"[ERROR] Quick scalp execution: {e}")

    # NEW: Scalp Position Management
    async def _manage_scalp_position(self, current_price: float):
        """Manage quick scalp positions with tight TP/SL"""
        if self._current_position is None or not self._current_position.get('is_scalp', False):
            return
            
        pos = self._current_position
        current_time = datetime.now(timezone.utc)
        hold_duration = (current_time - pos['scalp_start_time']).total_seconds()
        
        # Force close if max duration reached
        if hold_duration >= self.cfg.max_scalp_duration:
            await self._close_position(pos['side'], current_price, "SCALP_TIMEOUT")
            return
            
        # Check TP/SL
        if pos['side'] == 'LONG':
            if current_price >= pos['tp_price']:
                await self._close_position('LONG', current_price, "SCALP_TP")
            elif current_price <= pos['sl_price']:
                await self._close_position('LONG', current_price, "SCALP_SL")
        else:  # SHORT
            if current_price <= pos['tp_price']:
                await self._close_position('SHORT', current_price, "SCALP_TP")
            elif current_price >= pos['sl_price']:
                await self._close_position('SHORT', current_price, "SCALP_SL")

    # MODIFIED: Enhanced Emergency Exit Check dengan Tick Support
    async def _emergency_exit_check(self, price: float):
        """Enhanced emergency exit with scalp position support"""
        if self._current_position is None:
            return

        pos = self._current_position
        
        # Handle scalp positions separately
        if pos.get('is_scalp', False):
            await self._manage_scalp_position(price)
            return

        # Existing emergency exit logic for normal positions
        entry_price = pos['entry_price']
        tp = pos['tp_price']
        sl = pos['sl_price']
        entry_time = pos['entry_time']
        side = pos['side']

        elapsed_sec = (datetime.now(timezone.utc) - entry_time).total_seconds()
        profit_pct = self.calc_profit_percent(entry_price, side, price, leverage=self.cfg.leverage)
        
        # Existing guard exit conditions...
        if elapsed_sec >= self.cfg.max_hold_guard_sec and profit_pct >= self.cfg.guard_profit_trigger:
            print(f"[GUARD EXIT] TP {side} profit {profit_pct*100:.2f}% after {elapsed_sec/60:.1f} min")
            await self._close_position(side, price, "GUARD_PROFIT_LOCK")
            return

        if elapsed_sec >= self.cfg.max_hold_guard_sec2 and profit_pct >= self.cfg.guard_profit_trigger2:
            print(f"[GUARD EXIT] TP {side} profit {profit_pct*100:.2f}% after {elapsed_sec/60:.1f} min")
            await self._close_position(side, price, "GUARD_PROFIT_LOCK")
            return
        
        if elapsed_sec >= self.cfg.max_hold_guard_sec3 and profit_pct >= self.cfg.guard_profit_trigger3:
            print(f"[GUARD EXIT] TP {side} profit {profit_pct*100:.2f}% after {elapsed_sec/60:.1f} min")
            await self._close_position(side, price, "GUARD_PROFIT_LOCK")
            return

        if elapsed_sec >= self.cfg.max_hold_guard_sl_sec and profit_pct <= self.cfg.guard_loss_trigger:
            print(f"[GUARD EXIT] SL {side} profit {profit_pct*100:.2f}% after {elapsed_sec/60:.1f} min")
            await self._close_position(side, price, "GUARD_STOP_LOST")
            return

        # Standard TP/SL check
        if side == "LONG":
            if price <= sl:
                await self._close_position("LONG", price, "EMERGENCY_SL")
            elif price >= tp:
                await self._close_position("LONG", price, "EMERGENCY_TP")
        elif side == "SHORT":
            if price >= sl:
                await self._close_position("SHORT", price, "EMERGENCY_SL")
            elif price <= tp:
                await self._close_position("SHORT", price, "EMERGENCY_TP")

    # NEW: Process Tick Data
    async def _process_tick_data(self, price: float, volume: float, is_buyer_maker: bool, timestamp: datetime):
        """Process incoming tick data for quick trading decisions"""
        
        # Store tick data
        self.tick_prices.append(price)
        self.tick_volumes.append(volume)
        self.tick_timestamps.append(timestamp)
        self.last_tick_price = price
        self.last_tick_time = timestamp
        
        # Mark analysis ready after collecting enough data
        if len(self.tick_prices) >= 10 and not self.tick_analysis_ready:
            self.tick_analysis_ready = True
            print("[TICK] Tick analysis ready - monitoring for quick entries")
        
        # Process tick data only if enabled and ready
        if self.cfg.enable_tick_trading and self.tick_analysis_ready:
            # Check for quick scalp entries
            if self._current_position is None and not self.trade_locked:
                tick_signal = self._check_tick_entry_signal(price, is_buyer_maker)
                if tick_signal:
                    print(f"⚡ [TICK SIGNAL] {tick_signal} detected at {price:.3f}")
                    await self._execute_quick_scalp(tick_signal, price)
            
            # Manage existing positions with tick data
            await self._emergency_exit_check(price)

    # MODIFIED: Enhanced Start Method dengan Tick Stream
    async def start(self):
        self.client = await AsyncClient.create(self.cfg.api_key, self.cfg.api_secret, testnet=self.cfg.use_testnet)
        await self._load_symbol_precision()
        # await self.test_precision()
        try:
            await self.client.futures_change_leverage(symbol=self.cfg.pair, leverage=int(self.cfg.leverage))
            try:
                await self.client.futures_change_margin_type(symbol=self.cfg.pair, marginType=self.cfg.margin_type)
            except Exception:
                pass
        except Exception as e:
            print("[WARN] set leverage/margin:", e)

        print("[INIT] Loading historical candle data...")

        try:
            # Load 100 candle historis (8+ jam data 5m)
            historical_candles = await self.client.futures_klines(
                symbol=self.cfg.pair,
                interval=self.cfg.interval,
                limit=100
            )
            
            # Process dan simpan candle historis
            for candle in historical_candles:
                ts = pd.to_datetime(candle[0], unit='ms', utc=True)
                o, h, l, c = map(float, candle[1:5])
                v = float(candle[5])
                self._append_candle(ts, o, h, l, c, v)
            
            print(f"[INIT] Successfully loaded {len(self.candles)} historical candles")
            print(f"[INIT] Latest candle time: {self.candles.index[-1]}")

            # Hitung indikator teknikal berdasarkan data historis
            self.volatility_ratio = compute_volatility_ratio(self.candles, self.cfg.atr_period)
            print(f"[INIT] Initial volatility ratio: {self.volatility_ratio:.4f}")  
        except Exception as e:
            print(f"[ERROR] Failed to load historical data: {e}")
        
        self.bm = BinanceSocketManager(self.client)
        print(f"[INFO] Starting enhanced live-dual-entry bot for {self.cfg.pair}")
        print(f"[TICK] Tick trading: {'ENABLED' if self.cfg.enable_tick_trading else 'DISABLED'}")
        
        # Start periodic tasks
        asyncio.create_task(self.periodic_order_checks())
        
        # Run both kline and tick streams concurrently
        await asyncio.gather(
            self._run_kline_stream(),
            self._run_tick_stream()
        )

    # NEW: Tick Stream Handler
    async def _run_tick_stream(self):
        """Handle tick data stream"""
        print("[TICK] Starting tick data stream...")
        async with self.bm.trade_socket(self.cfg.pair) as stream:
            while True:
                try:
                    res = await stream.recv()
                    
                    # Extract tick data
                    price = float(res['p'])
                    volume = float(res['q'])
                    is_buyer_maker = res['m']  # True if seller is aggressive (bearish)
                    timestamp = pd.to_datetime(res['T'], unit='ms', utc=True)
                    
                    # Process tick data
                    await self._process_tick_data(price, volume, is_buyer_maker, timestamp)
                    
                except Exception as e:
                    print(f"[ERROR] tick stream: {e}")
                    await asyncio.sleep(1)

    # NEW: Kline Stream Handler (dipisah dari main loop)
    async def _run_kline_stream(self):
        """Handle kline data stream for candle-based analysis"""
        print("[CANDLE] Starting kline data stream...")
        async with self.bm.kline_futures_socket(self.cfg.pair, interval=self.cfg.interval) as stream:
            while True:
                try:
                    res = await stream.recv()
                    k = res.get('k', {})
                    is_closed = k.get('x', False)
                    ts = pd.to_datetime(k.get('t'), unit='ms', utc=True)
                    o, h, l, c, v = map(float, (k.get('o'), k.get('h'), k.get('l'), k.get('c'), k.get('v')))
                    
                    self._append_candle(ts, o, h, l, c, v)
                    self.volatility_ratio = compute_volatility_ratio(self.candles, self.cfg.atr_period)

                    await self._check_daily_reset()
                    await self._update_daily_profit()

                    if not self.trade_locked and self.daily_realized_pct >= self.cfg.daily_profit_lock_pct:
                        print(f"[LOCK] Target harian {self.daily_realized_pct:.2f}% tercapai → locking trading.")
                        await self.client.futures_cancel_all_open_orders(symbol=self.cfg.pair)
                        self.trade_locked = True

                    if is_closed:
                        atr_series = compute_atr_from_df(self.candles, self.cfg.atr_period)
                        current_atr = atr_series.iat[-1] if len(atr_series) >= self.cfg.atr_period else np.nan

                        if not np.isnan(current_atr) and current_atr >= self.cfg.min_atr:
                            self._create_watch(current_atr)

                        await self._process_watches()
                        await self._process_current_position()
                        # await self._cancel_misaligned_orders()
                        await self._check_pending_orders()

                except Exception as e:
                    print(f"[ERROR] kline stream: {e}")
                    await asyncio.sleep(1)

    def calculate_order_quality(self, order: Dict) -> float:
        """Calculate a quality score for order prioritization"""
        # Factors to consider:
        # 1. Distance from current price (closer is better)
        # 2. Volatility multiplier (higher volatility might be better for scalping)
        # 3. Time since signal (newer signals might be better)
        
        current_price = self.candles['close'].iloc[-1] if not self.candles.empty else 0
        price_distance = abs(order['entry_price'] - current_price) / current_price if current_price else 1
        
        # Lower distance = higher quality
        distance_score = 1 - min(price_distance, 0.1) * 10  # Normalize to 0-1 range
        
        # Higher volatility multiplier = higher quality (for aggressive trading)
        volatility_score = order.get('vol_mult', 1.0)
        
        # Newer orders are generally better
        time_score = 1.0  # Can be based on order time if needed
        
        return distance_score * 0.5 + volatility_score * 0.3 + time_score * 0.2

    # MODIFIED: Enhanced Position Close dengan Trade Type Tracking
    async def _close_position(self, side: str, exit_price: float, reason: str = "EMERGENCY"):
        if self._current_position is None:
            return
        
        pos = self._current_position
        try:
            close_order = await self.client.futures_create_order(
                symbol=self.cfg.pair,
                side=SIDE_SELL if side == 'LONG' else SIDE_BUY,
                type=ORDER_TYPE_MARKET,
                reduceOnly=True,
                quantity=pos['qty']
            )

            # Calculate execution price
            exit_exec_price = 0.0
            exit_qty = 0.0
            if 'fills' in close_order and close_order['fills']:
                for fill in close_order['fills']:
                    exit_qty += float(fill['qty'])
                    exit_exec_price += float(fill['price']) * float(fill['qty'])
                exit_exec_price = exit_exec_price / exit_qty if exit_qty > 0 else exit_price
            else:
                exit_exec_price = exit_price

            # Apply slippage
            if pos['side'] == 'LONG':
                exit_exec_price *= 1 - self.cfg.slippage_pct
            else:
                exit_exec_price *= 1 + self.cfg.slippage_pct

            exit_exec_price = self._round_price(exit_exec_price)

            # Calculate PnL
            raw_pnl = pos['qty'] * ((exit_exec_price - pos['entry_price']) if pos['side'] == 'LONG' else (pos['entry_price'] - exit_exec_price))
            fees = self.cfg.fee_rate * self.balance
            net_pnl = raw_pnl - fees
            self.balance += net_pnl

            # Determine trade type for logging
            trade_type = "SCALP" if pos.get('is_scalp', False) else "SWING"

            # Log to Excel with trade type
            append_trade_excel(self.cfg.logfile, [
                self.cfg.pair,
                pos['entry_time'],
                pos['side'],
                pos['entry_price'],
                pos['qty'],
                pos.get('tp_price', 0),
                pos.get('sl_price', 0),
                datetime.now(timezone.utc),
                exit_exec_price,
                net_pnl,
                fees,
                reason,
                self.balance,
                pos.get('atr', 0),
                pos.get('vol_mult', 1.0),
                trade_type  # NEW: Add trade type
            ])

            print(f"[POSITION CLOSED] {pos['side']} {trade_type} exit={exit_exec_price:.3f} reason={reason} pnl={net_pnl:.6f} balance={self.balance:.4f}")
            
            # Cleanup
            self._current_position = None
            self.watches = [w for w in self.watches if w['expire_idx'] > len(self.candles)-1]
            
        except Exception as e:
            print("[ERROR] closing position:", e)
    
    async def manage_order_conflicts(self, new_order: Dict):
        """Handle multiple order requests based on priority"""
        # Strategy 1: Cancel oldest order if new order has better potential
        if len(self.pending_orders) >= 2:  # Max 2 pending orders
            # Cancel the oldest order
            oldest_order = min(self.pending_orders, key=lambda x: x['place_time'])
            try:
                await self.client.futures_cancel_order(
                    symbol=self.cfg.pair,
                    orderId=oldest_order['order_id']
                )
                self.pending_orders.remove(oldest_order)
                print(f"[ORDER CANCELED] Conflict resolution - removed oldest order {oldest_order['order_id']}")
            except Exception as e:
                print(f"[ERROR] canceling order: {e}")
        
        # Strategy 2: Compare order quality and keep the best one
        current_best = None
        if self.pending_orders:
            current_best = max(self.pending_orders, key=lambda x: x.get('quality_score', 0))
        
        new_order_quality = self.calculate_order_quality(new_order)
        
        if current_best and new_order_quality > current_best.get('quality_score', 0):
            # New order is better, cancel the current one
            try:
                await self.client.futures_cancel_order(
                    symbol=self.cfg.pair,
                    orderId=current_best['order_id']
                )
                self.pending_orders.remove(current_best)
                print(f"[ORDER CANCELED] Replaced with better order")
            except Exception as e:
                print(f"[ERROR] canceling order: {e}")

    async def _update_daily_profit(self):
        info = await self.client.futures_account_balance()
        # print(info)
        # cari row untuk USDT
        usdt_row = next((x for x in info if x['asset'] == 'USDT'), None)
        
        if usdt_row:
            current_equity = float(usdt_row['availableBalance'])

            # Jika belum di-set, anggap equity awal = equity sekarang
            if self.daily_start_equity is None:
                self.daily_start_equity = current_equity

            self.daily_realized_pct = (
                (current_equity - self.daily_start_equity) / self.daily_start_equity * 100
            )

            daily_loss_limit = -1.0  # 0.5% max daily loss
            if self.daily_realized_pct <= daily_loss_limit:
                print(f"🚨 DAILY LOSS LIMIT REACHED: {self.daily_realized_pct:.2f}%")
                print(f"🚨 STOPPING ALL TRADING FOR TODAY")
                self.trade_locked = True
                await self._force_close_all2()

                # print(f"[INFO] Profit harian: {self.daily_realized_pct:.2f}% "
                #     f"(Equity: {current_equity:.2f} USDT)")

        else:
            print("[WARN] USDT balance tidak ditemukan")
    
    async def _force_close_all2(self):
        print("[LOCK] Closing all open positions...")
        
        positions = await self.client.futures_position_information()
        symbols = set()  # untuk kumpulkan simbol

        for pos in positions:
            qty = float(pos['positionAmt'])
            symbol = pos['symbol']
            symbols.add(symbol)

            if qty != 0:
                side = 'SELL' if qty > 0 else 'BUY'
                await self.client.futures_create_order(
                    symbol=symbol,
                    side=side,
                    type='MARKET',
                    quantity=abs(qty)
                )
                print(f"[FORCE CLOSE] {symbol} {side} {abs(qty)} closed.")

        print("[LOCK] Canceling all pending orders...")
        for sym in symbols:
            try:
                await self.client.futures_cancel_all_open_orders(symbol=sym)
                print(f"[LOCK] All pending orders for {sym} canceled.")
            except Exception as e:
                print(f"[LOCK][WARN] Gagal cancel {sym}: {e}")
    
    async def check_order_status(self):
        """Periodically check status of all pending orders"""
        if not self.client:
            return
            
        for order in self.pending_orders[:]:
            try:
                # Check order status from exchange
                order_status = await self.client.futures_get_order(
                    symbol=self.cfg.pair,
                    orderId=order['order_id']
                )
                
                if order_status['status'] == 'FILLED':
                    # Order filled, move to active orders
                    print(f"[ORDER FILLED] {order['side']} {order['qty']} @ {order_status['avgPrice']}")
                    self.pending_orders.remove(order)
                    
                    self.active_orders.append({
                        **order,
                        'entry_price': float(order_status['avgPrice']),
                        'entry_time': datetime.now(timezone.utc),
                        'status': 'FILLED'
                    })

                    self._current_position = {
                        **order,
                        'entry_price': float(order_status['avgPrice']),
                        'entry_time': datetime.now(timezone.utc),
                        'status': 'FILLED'
                    }
                    
                    try:
                        # await self.client.futures_cancel_all_open_orders(symbol=self.cfg.pair)
                        # print("[CANCEL] Semua pending order limit/stop dibatalkan.")
                        open_orders = await self.client.futures_get_open_orders(symbol=self.cfg.pair)
                        for o in open_orders:
                            if o["orderId"] != order_status["orderId"]:  # jangan cancel order yg barusan filled
                                await self.client.futures_cancel_order(symbol=self.cfg.pair, orderId=o["orderId"])
                                print(f"[CANCEL] Pending order {o['orderId']} side={o['side']} qty={o['origQty']}")
                    except Exception as e:
                        print("[ERROR] Cancel open orders failed:", e)

                elif order_status['status'] == 'CANCELED' or order_status['status'] == 'EXPIRED':
                    # Order canceled or expired, remove from pending
                    print(f"[ORDER {order_status['status']}] {order['order_id']}")
                    self.pending_orders.remove(order)
                    
            except Exception as e:
                print(f"[ERROR] checking order status: {e}")
    
    def _round_price(self, price: float) -> float:
        """Round price to configured precision"""
        return round(price, self.cfg.price_precision)

    def _round_qty(self, qty: float) -> float:
        """Round quantity to configured precision"""
        return round(qty, self.cfg.qty_precision)
    
    async def periodic_order_checks(self):
        """Run order checks periodically"""
        while True:
            try:
                await asyncio.sleep(30)  # Check every 30 seconds
                await self.check_order_status()
                await self.check_order_timeouts()
            except Exception as e:
                print("[ERROR] in periodic order checks:", e)
    
    def calculate_proper_position_size(self, entry_price: float, sl_price: float) -> float:
        """Calculate position size based on risk percentage"""
        risk_amount = self.balance * self.cfg.risk_pct  # 0.8% risk per trade
        price_risk = abs(entry_price - sl_price)
        
        if price_risk == 0:
            return 0
            
        raw_qty = risk_amount / price_risk
        return self._round_qty(raw_qty)

    async def long_strategy(self, side: str, entry_price: float, tp_price: float, sl_price: float, atr_value: float, vol_mult: float, market_regime: str):
        """Master long strategy berdasarkan market regime"""
        
        if market_regime == "STRONG_UPTREND":
            # 🚀 TREND-FOLLOWING LONG
            print("🚀 UPTREND - Trend-following long dengan market order")
            qty = self.calculate_proper_position_size(entry_price, sl_price) * 1.0
            qty = adjust_qty(qty,self.cfg.price_precision)
            print(f"Qty: {qty}")
            await self._open_market_position(
                side, entry_price, tp_price, sl_price,
                atr_value, vol_mult, qty
            )
            
        elif market_regime == "SIDEWAYS":
            # 🔄 MEAN-REVERSION LONG  
            print("🔄 SIDEWAYS - Mean-reversion long dengan limit order")
            reduced_qty = self.calculate_proper_position_size(entry_price, sl_price) * 0.7
            reduced_qty = adjust_qty(reduced_qty,self.cfg.price_precision)
            print(f"Qty: {reduced_qty}")
            await self._place_limit_order(
                side, entry_price, tp_price, sl_price,
                atr_value, vol_mult, reduced_qty
            )
            
        elif market_regime == "MIXED":
            # ⚠️ CAUTIOUS LONG
            print("⚠️  MIXED - Cautious long dengan reduced size")
            reduced_qty = self.calculate_proper_position_size(entry_price, sl_price) * 0.5
            reduced_qty = adjust_qty(reduced_qty,self.cfg.price_precision)
            print(f"Qty: {reduced_qty}")
            await self._place_limit_order(
                side, entry_price, tp_price, sl_price,
                atr_value, vol_mult, reduced_qty
            )

    async def short_strategy(self, side: str, entry_price: float, tp_price: float, sl_price: float, atr_value: float, vol_mult: float, market_regime: str):
        """Master short strategy berdasarkan market regime"""
        
        if market_regime == "STRONG_DOWNTREND":
            # 🚀 TREND-FOLLOWING SHORT
            print("🚀 DOWNTREND - Trend-following short dengan market order")
            qty = self.calculate_proper_position_size(entry_price, sl_price) * 1.0
            qty = adjust_qty(qty,self.cfg.price_precision)
            await self._open_market_position(
                side, entry_price, tp_price, sl_price,
                atr_value, vol_mult, qty
            )
            
        elif market_regime == "SIDEWAYS":
            # 🔄 MEAN-REVERSION SHORT
            print("🔄 SIDEWAYS - Mean-reversion short dengan limit order") 
            reduced_qty = self.calculate_proper_position_size(entry_price, sl_price) * 0.7
            reduced_qty = adjust_qty(reduced_qty,self.cfg.price_precision)
            await self._place_limit_order(
                side, entry_price, tp_price, sl_price,
                atr_value, vol_mult, reduced_qty
            )
            
        elif market_regime == "MIXED":
            # ⚠️ CAUTIOUS SHORT
            print("⚠️  MIXED - Cautious short dengan reduced size")
            reduced_qty = self.calculate_proper_position_size(entry_price, sl_price) * 0.5
            reduced_qty = adjust_qty(reduced_qty,self.cfg.price_precision)
            await self._place_limit_order(
                side, entry_price, tp_price, sl_price,
                atr_value, vol_mult, reduced_qty
            )

    def calculate_directional_indicators(self):
        """Hitung Plus DI dan Minus DI menggunakan TA-Lib"""
        if len(self.candles) < 15:
            return 0, 0, 0
        
        high = self.candles['high'].values
        low = self.candles['low'].values  
        close = self.candles['close'].values
        
        # Calculate DI+ dan DI-
        plus_di = talib.PLUS_DI(high, low, close, timeperiod=14)
        minus_di = talib.MINUS_DI(high, low, close, timeperiod=14)
        adx = talib.ADX(high, low, close, timeperiod=14)
        
        # Ambil nilai terakhir
        current_plus_di = plus_di[-1] if len(plus_di) > 0 else 0
        current_minus_di = minus_di[-1] if len(minus_di) > 0 else 0
        current_adx = adx[-1] if len(adx) > 0 else 0
        
        return current_plus_di, current_minus_di, current_adx
    
    def calculate_ema_indicators(self):
        """Hitung EMA Fast (20) dan EMA Slow (50)"""
        if len(self.candles) < 50:
            return 0, 0
        
        close = self.candles['close'].values
        
        # EMA Fast (20 periode) - untuk short-term trend
        ema_fast = talib.EMA(close, timeperiod=20)
        
        # EMA Slow (50 periode) - untuk medium-term trend  
        ema_slow = talib.EMA(close, timeperiod=50)
        
        # EMA Long (200 periode) - untuk long-term trend (optional)
        ema_long = talib.EMA(close, timeperiod=200)
        
        current_ema_fast = ema_fast[-1] if len(ema_fast) > 0 else 0
        current_ema_slow = ema_slow[-1] if len(ema_slow) > 0 else 0
        current_ema_long = ema_long[-1] if len(ema_long) > 0 else 0
        
        return current_ema_fast, current_ema_slow, current_ema_long

    def enhanced_trend_detection(self) -> Dict:
        """Enhanced trend detection yang lebih akurat dan konsisten"""
        if len(self.candles) < 30:  # Reduced from 50 to 30
            return {"regime": "MIXED", "confidence": 0.3, "trend_direction": "NEUTRAL"}
        
        close = self.candles['close'].values
        high = self.candles['high'].values
        low = self.candles['low'].values
        
        # Dapatkan indikator dengan error handling
        try:
            plus_di, minus_di, adx = self.calculate_directional_indicators()
            ema_fast, ema_slow, ema_long = self.calculate_ema_indicators()
            rsi = talib.RSI(close, timeperiod=14)[-1] if len(close) >= 14 else 50
        except Exception as e:
            print(f"[WARN] Indicator calculation error: {e}")
            return {"regime": "MIXED", "confidence": 0.3, "trend_direction": "NEUTRAL"}
        
        # Price structure analysis - lebih robust
        if len(high) >= 10 and len(low) >= 10:
            recent_highs = high[-10:]  # Reduced from 20 to 10
            recent_lows = low[-10:]
            higher_highs = all(recent_highs[i] > recent_highs[i-1] for i in range(1, len(recent_highs)))
            lower_lows = all(recent_lows[i] < recent_lows[i-1] for i in range(1, len(recent_lows)))
        else:
            higher_highs = lower_lows = False
        
        # Scoring system yang lebih realistis
        trend_score = 0
        max_score = 8  # Reduced from 10
        
        # 1. ADX Strength (2 points)
        if adx > 25: 
            trend_score += 2
        elif adx > 18:  # Reduced threshold
            trend_score += 1
        
        # 2. EMA Alignment (2 points) - simplified
        if ema_fast > ema_slow and ema_slow > ema_long:
            trend_score += 2  # Strong uptrend
        elif ema_fast < ema_slow and ema_slow < ema_long:
            trend_score += 2  # Strong downtrend
        elif (ema_fast > ema_slow) != (ema_slow > ema_long):  # Mixed alignment
            trend_score += 0  # No points for mixed
        
        # 3. DI Alignment (2 points)
        di_diff = plus_di - minus_di
        if abs(di_diff) > 15:  # Significant DI difference
            trend_score += 2
        elif abs(di_diff) > 8:  # Moderate DI difference
            trend_score += 1
        
        # 4. Price Structure (2 points)
        if higher_highs and not lower_lows:
            trend_score += 2  # Uptrend structure
        elif lower_lows and not higher_highs:
            trend_score += 2  # Downtrend structure
        
        confidence = trend_score / max_score
        
        # Tentukan regime dengan threshold yang lebih reasonable
        if trend_score >= 5 and confidence > 0.6:  # Reduced thresholds
            if plus_di > minus_di:
                return {"regime": "STRONG_TREND", "confidence": confidence, "trend_direction": "UP"}
            else:
                return {"regime": "STRONG_TREND", "confidence": confidence, "trend_direction": "DOWN"}
        elif trend_score <= 2 and confidence < 0.3:  # Reduced thresholds
            return {"regime": "SIDEWAYS", "confidence": confidence, "trend_direction": "NEUTRAL"}
        else:
            return {"regime": "MIXED", "confidence": confidence, "trend_direction": "NEUTRAL"}
    
    def enhanced_sideways_detection(self) -> bool:
        """Enhanced sideways detection dengan kondisi yang lebih realistis"""
        if len(self.candles) < 20:
            return False
        
        try:
            close = self.candles['close'].values
            high = self.candles['high'].values
            low = self.candles['low'].values
            
            # 1. ADX rendah (di bawah 22 - increased threshold)
            plus_di, minus_di, adx = self.calculate_directional_indicators()
            if adx > 22:  # Increased from 20
                return False
            
            # 2. EMA berdekatan (threshold increased)
            ema_fast, ema_slow, _ = self.calculate_ema_indicators()
            ema_diff_pct = abs(ema_fast - ema_slow) / close[-1]
            if ema_diff_pct > 0.015:  # Increased from 0.01
                return False
            
            # 3. Price range sempit (menggunakan ATR - threshold increased)
            atr = talib.ATR(high, low, close, timeperiod=14)[-1]
            atr_pct = atr / close[-1]
            if atr_pct > 0.008:  # Increased from 0.005
                return False
            
            # 4. RSI netral (tambahan kondisi baru)
            rsi = talib.RSI(close, timeperiod=14)[-1] if len(close) >= 14 else 50
            if rsi < 40 or rsi > 60:  # RSI di luar range netral
                return False
            
            # Jika semua kondisi terpenuhi, maka sideways
            print(f"[SIDEWAYS DETECTED] ADX:{adx:.1f}, EMA_diff:{ema_diff_pct:.3f}, ATR%:{atr_pct:.3f}, RSI:{rsi:.1f}")
            return True
            
        except Exception as e:
            print(f"[ERROR] Sideways detection: {e}")
            return False
    
    def _validate_candle_data(self) -> bool:
        """Validasi kualitas dan kelengkapan data candle"""
        if len(self.candles) < 20:
            print("[WARN] Not enough candle data for analysis")
            return False
        
        # Cek ada NaN values
        if self.candles.isnull().any().any():
            print("[WARN] Candle data contains NaN values")
            return False
        
        # Cek volume (jangan terlalu rendah)
        recent_volume = self.candles['volume'].tail(5).mean()
        if recent_volume < 1000:  # Minimum volume threshold
            print(f"[WARN] Low volume: {recent_volume:.0f}")
            return False
        
        return True
    
    def get_daily_size_multiplier(self) -> float:
        """Reduce position size as we approach daily target"""
        if self.daily_realized_pct >= 1.0:
            return 0.0  # Stop at 1%
        elif self.daily_realized_pct >= 0.8:
            return 0.3  # 30% size near target
        elif self.daily_realized_pct >= 0.5:
            return 0.6  # 60% size at halfway
        else:
            return 1.0  # Full size
    
    async def _cancel_misaligned_orders(self,reverse=False):
        """
        Membatalkan limit order yang arah-nya sudah tidak sama dengan sinyal aktif terakhir.
        """
        signal_side = self._current_signal_side
        if reverse :
            signal_side = "LONG" if side == "SHORT" else "SHORT"

        if signal_side is None:
            return

        try:
            open_orders = await self.client.futures_get_open_orders(symbol=self.cfg.pair)
            for od in open_orders:
                side = od.get('side')  # 'BUY' atau 'SELL'
                order_side = 'LONG' if side == 'BUY' else 'SHORT'

                if order_side != side:
                    await self.client.futures_cancel_order(
                        symbol=self.cfg.pair,
                        orderId=od['orderId']
                    )
                    print(f"[CANCEL] {order_side} limit dibatalkan karena sinyal berubah → {signal_side}")
        except Exception as e:
            print(f"[ERROR] gagal cancel misaligned orders: {e}")
    
    async def _check_daily_reset(self):
        now = datetime.now().hour

        if now == self.cfg.daily_reset_hour:         # selalu cek jam reset
            # Hindari reset berulang di jam yang sama
            if getattr(self, "_last_reset_day", None) != datetime.now().day:
                info = await self.client.futures_account_balance()
                usdt_row = next((x for x in info if x['asset'] == 'USDT'), None)

                if usdt_row:
                    self.daily_start_equity = float(usdt_row['balance'])
                    self.daily_realized_pct = 0
                    self.trade_locked = False         # buka kunci untuk hari baru
                    self._last_reset_day = datetime.now().day

                    print(f"[RESET] Daily profit lock reset for new trading day. "
                        f"Start equity = {self.daily_start_equity:.2f} USDT")
                else:
                    print("[WARN] USDT balance tidak ditemukan saat reset!")
    
    def _detect_trend_mode(self):
        close = self.candles['close'].values
        atr = talib.ATR(
            self.candles['high'].values,
            self.candles['low'].values,
            close,
            timeperiod=14
        )[-1]

        adx = talib.ADX(
            self.candles['high'].values,
            self.candles['low'].values,
            close,
            timeperiod=14
        )[-1]

        # aturan sederhana: trend kalau ADX > 25 dan jarak harga dari MA50 > ATR*1.5
        ma50 = talib.SMA(close, timeperiod=50)[-1]
        dist = abs(close[-1] - ma50)

        if adx > 20 and dist > 1.2 * atr:
            return "trend"
        return "sideways"

    def calc_profit_percent(self,entry_price: float, side: str, latest_price: float,leverage = 1) -> float:
        if side.upper() == "LONG":
            raw =  (latest_price - entry_price) / entry_price
        elif side.upper() == "SHORT":
            raw = (entry_price - latest_price) / entry_price
        else:
            raise ValueError("side harus 'LONG' atau 'SHORT'")
        
        return raw * leverage

    def _append_candle(self, ts, o, h, l, c, v):
        row = pd.DataFrame([[o, h, l, c, v]], index=[ts], columns=['open', 'high', 'low', 'close', 'volume'])
        if self.candles.empty:
            self.candles = row
        else:
            self.candles = pd.concat([self.candles, row])
        if len(self.candles) > self.cfg.candles_buffer:
            self.candles = self.candles.iloc[-self.cfg.candles_buffer:]
    
    def _create_watch(self, atr_value):
        if self._current_position is not None:
            return
        last_close = self.candles['close'].iat[-1]
        
        volatility_mult = 1.0
        if self.cfg.adaptive_tp_sl and self.volatility_ratio > 0.005:
            volatility_mult = 1.2
        elif self.cfg.adaptive_tp_sl and self.volatility_ratio < 0.002:
            volatility_mult = 0.8
            
        watch = {
            "start_idx": len(self.candles) - 1,
            "expire_idx": len(self.candles) - 1 + self.cfg.monitor_candles,
            "long_level": self._round_price(last_close + atr_value * self.cfg.level_mult * volatility_mult),
            "short_level": self._round_price(last_close - atr_value * self.cfg.level_mult * volatility_mult),
            "atr": atr_value,
            "trigger_time": self.candles.index[-1],
            "volatility_mult": volatility_mult
        }
        self.watches.append(watch)
        print(f"[WATCH CREATED] {watch['trigger_time']} ATR={atr_value:.6f} long={watch['long_level']:.3f} short={watch['short_level']:.3f} vol_mult={volatility_mult:.2f}")


    async def _process_watches(self, reverse=False):
        if self._current_position is not None:
            return

         # VALIDASI DATA CANDLE SEBELUM ANALISIS
        if not self._validate_candle_data():
            print("[SKIP] Candle data validation failed - skipping analysis")
            return
        
        # ENHANCED MARKET DETECTION
        market_analysis = self.enhanced_trend_detection()
        regime = market_analysis['regime']
        confidence = market_analysis['confidence']
        trend_direction = market_analysis['trend_direction']

        is_sideways = self.enhanced_sideways_detection()
        if is_sideways and regime != "SIDEWAYS":
            print(f"[ADJUST] Overriding regime to SIDEWAYS based on detailed analysis")
            regime = "SIDEWAYS"
            confidence = 0.8

        print(f"[MARKET] Regime: {regime}, Direction: {trend_direction}, Confidence: {confidence:.2f}")

        size_multiplier = self.get_daily_size_multiplier()
        if size_multiplier == 0:
            print("🎯 Daily target reached - skipping new entries")
            return

        latest_idx = len(self.candles) - 1
        candle_high = self.candles['high'].iat[-1]
        candle_low = self.candles['low'].iat[-1]
        candle_close = self.candles['close'].iat[-1]

        mode = self._detect_trend_mode()
        print(f"[MODE] Market mode: {mode}")

        new_watches = []
        for w in self.watches:
            if latest_idx < w['start_idx']:
                new_watches.append(w)
                continue

            triggered = False
            side = None
            entry_price = None
            tp_price = None
            sl_price = None

            # --- DETEKSI KONDISI ENTRY ---
            if self.cfg.require_confirmation:
                long_condition = candle_close >= w['long_level']
                short_condition = candle_close <= w['short_level']
            else:
                long_condition = candle_high >= w['long_level']
                short_condition = candle_low <= w['short_level']

            # --- HITUNG TP/SL AWAL ---
            if long_condition:
                triggered = True
                side = 'LONG'
                self._current_signal_side = 'LONG'
                entry_price = w['long_level']
                tp_mult = self.cfg.tp_atr_mult * w['volatility_mult']
                sl_mult = self.cfg.sl_atr_mult * w['volatility_mult']
                tp_price = self._round_price(entry_price + w['atr'] * tp_mult)
                sl_price = self._round_price(entry_price - w['atr'] * sl_mult)

            elif short_condition:
                triggered = True
                side = 'SHORT'
                self._current_signal_side = 'SHORT'
                entry_price = w['short_level']
                tp_mult = self.cfg.tp_atr_mult * w['volatility_mult']
                sl_mult = self.cfg.sl_atr_mult * w['volatility_mult']
                tp_price = self._round_price(entry_price - w['atr'] * tp_mult)
                sl_price = self._round_price(entry_price + w['atr'] * sl_mult)

            # =====================
            #      JIKA TRIGGER
            # =====================
            if triggered:

                # Tentukan strategi berdasarkan regime dan trend direction
                if regime == "STRONG_TREND":
                    if trend_direction == "UP" and side == "LONG":
                        await self.long_strategy(side, entry_price, tp_price, sl_price, w['atr'], w['volatility_mult'], "STRONG_UPTREND")
                    elif trend_direction == "DOWN" and side == "SHORT":  
                        await self.short_strategy(side, entry_price, tp_price, sl_price, w['atr'], w['volatility_mult'], "STRONG_DOWNTREND")
                    else:
                        print(f"🚫 Trend direction {trend_direction} tidak match dengan signal {side}")
                        
                elif regime == "SIDEWAYS":
                    if side == "LONG":
                        await self.long_strategy(side, entry_price, tp_price, sl_price, w['atr'], w['volatility_mult'], "SIDEWAYS")
                    else:
                        await self.short_strategy(side, entry_price, tp_price, sl_price, w['atr'], w['volatility_mult'], "SIDEWAYS")
                        
                else:  # MIXED
                    if side == "LONG":
                        await self.long_strategy(side, entry_price, tp_price, sl_price, w['atr'], w['volatility_mult'], "MIXED")
                    else:
                        await self.short_strategy(side, entry_price, tp_price, sl_price, w['atr'], w['volatility_mult'], "MIXED")

            else:
                # Masih menunggu expire
                if latest_idx < w['expire_idx']:
                    new_watches.append(w)

        self.watches = new_watches
    
    async def _place_limit_order(self, side: str, entry_price: float, tp_price: float, sl_price: float, atr_value: float, vol_mult: float, qty: float):
        # qty = self._round_qty(1.0)  # 1 AVAX
        # qty = self.calculate_proper_position_size(entry_price, sl_price)
        if qty <= 0:
            print(f"[SKIP] Quantity too small: {qty}")
            return

        if side == 'LONG':
            order_price = entry_price * (1 - 0.0002)  # 0.05% below trigger
        else:
            order_price = entry_price * (1 + 0.0002)  # 0.05% above trigger

        # Round prices
        order_price = self._round_price(order_price)
        entry_price = self._round_price(entry_price)
        tp_price = self._round_price(tp_price)
        sl_price = self._round_price(sl_price)

        risk_per_trade = abs(entry_price - sl_price) * qty
        risk_percentage = (risk_per_trade / self.balance) * 100
        print(f"[DEBUG] Actual risk: {risk_percentage:.2f}% per trade")

        for po in list(self.pending_orders):
            if po['side'] == side:
                try:
                    # Ambil detail order dari exchange untuk cek filled quantity
                    order_info = await self.client.futures_get_order(
                        symbol=self.cfg.pair,
                        orderId=po['order_id']
                    )
                    filled = float(order_info.get("executedQty", 0))

                    if filled == 0:
                        # Aman untuk cancel
                        await self.client.futures_cancel_order(
                            symbol=self.cfg.pair,
                            orderId=po['order_id']
                        )
                        print(f"[REPLACE] Cancel order lama {side} @ {po['order_price']}")
                        self.pending_orders.remove(po)
                    else:
                        # Sudah terisi sebagian → jangan di-cancel
                        print(f"[SKIP] Order lama {side} @ {po['order_price']} "
                            f"sudah terisi sebagian ({filled} filled). Tidak di-cancel.")
                except Exception as e:
                    print(f"[WARN] Gagal cek/cancel order {po['order_id']}: {e}")

        if len(self.pending_orders) >= 2:  # Maximum 2 pending orders
            await self.manage_order_conflicts({
                'side': side,
                'entry_price': entry_price,
                'tp_price': tp_price,
                'sl_price': sl_price,
                'atr': atr_value,
                'vol_mult': vol_mult,
                'qty': qty
            })
            await asyncio.sleep(0.2)
        try:
            # Place the order
            order = await self.client.futures_create_order(
                symbol=self.cfg.pair,
                side=SIDE_BUY if side == 'LONG' else SIDE_SELL,
                type=ORDER_TYPE_LIMIT,
                timeInForce=TIME_IN_FORCE_GTC,
                quantity=qty,
                price=order_price
            )
            
            # Add to pending orders with additional metadata
            order_data = {
                "order_id": order['orderId'],
                "side": side,
                "entry_price": entry_price,
                "order_price": order_price,
                "qty": qty,
                "tp_price": tp_price,
                "sl_price": sl_price,
                "atr": atr_value,
                "vol_mult": vol_mult,
                "place_time": datetime.now(timezone.utc),
                "expiry_time": datetime.now(timezone.utc) + timedelta(seconds=self.order_timeout),
                "quality_score": self.calculate_order_quality({
                    'entry_price': entry_price,
                    'vol_mult': vol_mult
                })
            }
            
            self.pending_orders.append(order_data)
            # self.watch['order_id'] = order['orderId']
            print(f"[LIMIT ORDER PLACED] {side} {qty} @ {order_price:.3f} (orderId: {order['orderId']})")
            
        except Exception as e:
            print("[ERROR] place limit order:", e)

    async def _load_symbol_precision(self):
        """Load precision rules from Binance for AVAXUSDT"""
        try:
            exchange_info = await self.client.futures_exchange_info()
            for symbol_info in exchange_info['symbols']:
                if symbol_info['symbol'] == self.cfg.pair:
                    # Get LOT_SIZE filter for quantity
                    for f in symbol_info['filters']:
                        if f['filterType'] == 'LOT_SIZE':
                            step_size = float(f['stepSize'])
                            # Determine quantity precision from stepSize
                            if step_size == 1.0:
                                self.cfg.qty_precision = 0
                            elif step_size == 0.1:
                                self.cfg.qty_precision = 1
                            elif step_size == 0.01:
                                self.cfg.qty_precision = 2
                            elif step_size == 0.001:
                                self.cfg.qty_precision = 3
                            # print(f"[PRECISION] Step size: {step_size}, Qty precision: {self.cfg.qty_precision}")
                        
                        if f['filterType'] == 'PRICE_FILTER':
                            tick_size = float(f['tickSize'])
                            # Determine price precision from tickSize
                            if tick_size == 1.0:
                                self.cfg.price_precision = 0
                            elif tick_size == 0.1:
                                self.cfg.price_precision = 1
                            elif tick_size == 0.01:
                                self.cfg.price_precision = 2
                            elif tick_size == 0.001:
                                self.cfg.price_precision = 3
                            # print(f"[PRECISION] Tick size: {tick_size}, Price precision: {self.cfg.price_precision}")
                    
                    # print(f"[PRECISION] Final - Qty: {self.cfg.qty_precision}, Price: {self.cfg.price_precision}")
                    break
        except Exception as e:
            print(f"[ERROR] Loading symbol precision: {e}")
            # Fallback to default precision for AVAX
            self.cfg.qty_precision = 1
            self.cfg.price_precision = 3
    
    async def _open_market_position(self, side: str, entry_price: float, tp_price: float, sl_price: float, atr_value: float, vol_mult: float, qty: float):
        # qty = self._round_qty(1.0)  # 1 AVAX
        # qty = self.calculate_proper_position_size(entry_price, sl_price)

        if qty <= 0:
            print(f"[SKIP] Quantity too small: {qty}")
            return
        # Round prices
        entry_price = self._round_price(entry_price)
        tp_price = self._round_price(tp_price)
        sl_price = self._round_price(sl_price)

        risk_per_trade = abs(entry_price - sl_price) * qty
        risk_percentage = (risk_per_trade / self.balance) * 100
        print(f"[DEBUG] Actual risk: {risk_percentage:.2f}% per trade")
        

        try:
            if side == 'LONG':
                exec_price = entry_price * (1 + self.cfg.slippage_pct)
            else:
                exec_price = entry_price * (1 - self.cfg.slippage_pct)
                
            order = await self.client.futures_create_order(
                symbol=self.cfg.pair,
                side=SIDE_BUY if side == 'LONG' else SIDE_SELL,
                type=ORDER_TYPE_MARKET,
                quantity=qty
            )
            
            exec_qty = 0.0
            avg_price = 0.0
            if 'fills' in order and order['fills']:
                for fill in order['fills']:
                    exec_qty += float(fill['qty'])
                    avg_price += float(fill['price']) * float(fill['qty'])
                avg_price = avg_price / exec_qty if exec_qty > 0 else exec_price
            else:
                avg_price = exec_price
                exec_qty = qty

            avg_price = self._round_price(avg_price)
            print(f"[POSITION OPENED] {side} {exec_qty} @ {avg_price:.3f}")

            self.active_orders.append({
                        **order,
                        'entry_price': float(avg_price),
                        'entry_time': datetime.now(timezone.utc),
                        'status': 'FILLED'
                    })
            self._current_position = {
                        **order,
                        "entry_price": avg_price,
                        "entry_time": datetime.now(timezone.utc),
                        'status': 'FILLED'
                    }
        
        except Exception as e:
            print("[ERROR] open position:", e)
    
    async def _check_pending_orders(self):
        now = datetime.now(timezone.utc)
        for order in self.pending_orders[:]:
            if now >= order['expiry_time']:
                try:
                    await self.client.futures_cancel_order(
                        symbol=self.cfg.pair, 
                        orderId=order['order_id']
                    )
                    print(f"[ORDER CANCELED] {order['order_id']} (expired)")
                    self.pending_orders.remove(order)
                except Exception as e:
                    print(f"[ERROR] cancel order {order['order_id']}:", e)
    
    async def _process_current_position(self):
        if self._current_position is None:
            return

        pos = self._current_position
        latest_candle = self.candles.iloc[-1]
        now = self.candles.index[-1]
        elapsed_sec = (now.to_pydatetime() - pos['entry_time']).total_seconds()

        exit_reason = None
        exit_price = None
       
        if elapsed_sec >= self.cfg.min_hold_sec:
            calc_profit_percent = self.calc_profit_percent(
                pos['entry_price'],
                pos['side'],
                latest_candle['close'],
                self.cfg.leverage
            )
            print(f"Profit% (Binance style): {calc_profit_percent*100:.2f}%")

            if pos['side'] == 'LONG':
                if latest_candle['low'] <= pos['entry_price'] * (1 - 0.038):  # -3.8%
                    exit_price = latest_candle['low']
                    exit_reason = f"Quick SL intrabar"
            else:
                if latest_candle['high'] >= pos['entry_price'] * (1 + 0.038):  # -3.8%
                    exit_price = latest_candle['high']
                    exit_reason = f"Quick SL intrabar"

            if exit_price is None:
                if calc_profit_percent >= 0.038:  # +3.8% atau lebih
                    exit_price = latest_candle['close']
                    exit_reason = f"Quick TP (close)"

        if exit_price is not None:
            try:
                side = SIDE_SELL if pos['side'] == 'LONG' else SIDE_BUY
                close_order = await self.client.futures_create_order(
                    symbol=self.cfg.pair,
                    side=side,
                    type=ORDER_TYPE_MARKET,
                    reduceOnly=True,
                    quantity=pos['qty']
                )
                
                exit_exec_price = 0.0
                exit_qty = 0.0
                self._current_position = None
                # self.watches.clear()
                if 'fills' in close_order and close_order['fills']:
                    for fill in close_order['fills']:
                        exit_qty += float(fill['qty'])
                        exit_exec_price += float(fill['price']) * float(fill['qty'])
                    exit_exec_price = exit_exec_price / exit_qty if exit_qty > 0 else exit_price
                else:
                    exit_exec_price = exit_price
                    
                if pos['side'] == 'LONG':
                    exit_exec_price = exit_exec_price * (1 - self.cfg.slippage_pct)
                else:
                    exit_exec_price = exit_exec_price * (1 + self.cfg.slippage_pct)

                exit_exec_price = self._round_price(exit_exec_price)

                if pos['side'] == 'LONG':
                    raw_pnl = pos['qty'] * (exit_exec_price - pos['entry_price'])
                else:
                    raw_pnl = pos['qty'] * (pos['entry_price'] - exit_exec_price)

                fees = self.cfg.fee_rate * self.balance
                net_pnl = raw_pnl - fees

                self.balance += net_pnl

                append_trade_excel(self.cfg.logfile, [
                    self.cfg.pair,
                    pos['entry_time'],
                    pos['side'],
                    pos['entry_price'],
                    pos['qty'],
                    pos['tp_price'],
                    pos['sl_price'],
                    datetime.now(timezone.utc),
                    exit_exec_price,
                    net_pnl,
                    fees,
                    exit_reason,
                    self.balance,
                    pos['atr'],
                    pos['vol_mult']
                ])

                print(f"[POSITION CLOSED] {pos['side']} exit={exit_exec_price:.3f} reason={exit_reason} pnl={net_pnl:.6f} fees={fees:.6f} balance={self.balance:.4f}")
                self._current_position = None
                self.watches = [w for w in self.watches if w['expire_idx'] > len(self.candles)-1]
                
            except Exception as e:
                print("[ERROR] closing position:", e)
                self._current_position = None
                self.watches = [w for w in self.watches if w['expire_idx'] > len(self.candles)-1]


    async def check_order_timeouts(self):
        """Cancel orders that have been pending too long"""
        now = datetime.now(timezone.utc)
        for order in self.pending_orders[:]:
            if (now - order['place_time']).total_seconds() > self.order_timeout:
                try:
                    await self.client.futures_cancel_order(
                        symbol=self.cfg.pair,
                        orderId=order['order_id']
                    )
                    self.pending_orders.remove(order)
                    print(f"[ORDER CANCELED] Timeout - {order['order_id']}")
                except Exception as e:
                    print(f"[ERROR] canceling timed out order: {e}")
    
    async def test_precision(self):
        test_price = 30.225123
        test_qty = 1.234567
        
        rounded_price = self._round_price(test_price)
        rounded_qty = self._round_qty(test_qty)
        
        print(f"[PRECISION TEST] Price: {test_price} -> {rounded_price}")
        print(f"[PRECISION TEST] Qty: {test_qty} -> {rounded_qty}")

# ---------------- Run ----------------  
if __name__ == "__main__":
    cfg = BotConfig()
    cfg.api_key = os.getenv('BINANCE_API_KEY')
    cfg.api_secret = os.getenv('BINANCE_API_SECRET')
    cfg.use_testnet = False
    
    # Existing config
    cfg.use_limit_orders = True
    cfg.require_confirmation = True
    cfg.adaptive_tp_sl = True
    cfg.slippage_pct = 0.001
    cfg.qty_precision = 1
    cfg.price_precision = 3
    
    # NEW: Tick trading config
    cfg.enable_tick_trading = True  # Enable tick-based quick trading
    cfg.tick_buffer_size = 50
    cfg.min_tick_volume = 1000.0
    cfg.tick_momentum_period = 10
    cfg.quick_tp_pct = 0.005  # 0.5% quick TP
    cfg.quick_sl_pct = 0.003  # 0.3% quick SL
    cfg.max_scalp_duration = 300  # 5 minutes
    cfg.tick_volume_threshold = 2.0

    init_excel(cfg.logfile)
    bot = ImprovedLiveDualEntryBot(cfg)

    loop = asyncio.get_event_loop()
    try:
        loop.run_until_complete(bot.start())
    except KeyboardInterrupt:
        print("[STOP] Interrupted by user. Trades saved to", cfg.logfile)
    finally:
        try:
            if bot.client:
                loop.run_until_complete(bot.client.close_connection())
        except Exception:
            pass