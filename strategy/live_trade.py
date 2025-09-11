# live_bot_dual_entry.py
"""
Live forward-test (paper trading) - Dual-entry ATR breakout
Logic matches dual_entry_backtest() used earlier:
 - ATR filter
 - long_level = close + atr*level_mult
 - short_level = close - atr*level_mult
 - monitor next N closed candles for breakout
 - TP/SL scaled by ATR multipliers
 - PnL computed similar to backtest (balance * leverage * return - fee_rate*balance)
 - All trades saved into trades_log.xlsx
This script uses Binance WebSocket and runs forever (async).
It DOES NOT place real orders.
"""
import asyncio
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import List, Dict

import numpy as np
import pandas as pd
from binance import AsyncClient, BinanceSocketManager
from openpyxl import Workbook, load_workbook

# ---------------- Config ----------------
@dataclass
class BotConfig:
    pair: str = "AVAXUSDT"
    interval: str = "1m"
    initial_balance: float = 20.0
    leverage: float = 10.0
    fee_rate: float = 0.0004
    min_atr: float = 0.0005
    atr_period: int = 14
    level_mult: float = 0.2    # long/short trigger offset: close +/- atr*level_mult
    tp_atr_mult: float = 0.9
    sl_atr_mult: float = 0.9
    monitor_candles: int = 3   # how many closed candles to watch for breakout
    candles_buffer: int = 500   # how many recent candles to keep (for ATR etc.)
    logfile: str = "trades_log.xlsx"

cfg = BotConfig()

# ---------------- Excel helpers ----------------
def init_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append([
            "pair", "watch_start", "trigger_side", "trigger_level", "atr_at_trigger",
            "entry_time", "entry_price", "exit_time", "exit_price", "pnl",
            "exit_reason", "balance_after"
        ])
        wb.save(path)

def append_trade_excel(path: str, row: List):
    wb = load_workbook(path)
    ws = wb["Trades"]

    clean_row = []
    for v in row:
        if isinstance(v, datetime):
            clean_row.append(v.replace(tzinfo=None))  # buang timezone
        else:
            clean_row.append(v)

    ws.append(clean_row)
    wb.save(path)
# ---------------- Technical helpers ----------------
def compute_atr_from_df(df: pd.DataFrame, period: int):
    # df must have high, low, close
    high_low = df['high'] - df['low']
    high_close = (df['high'] - df['close'].shift(1)).abs()
    low_close = (df['low'] - df['close'].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    return atr

# ---------------- Watch structure ----------------
# when a "watch" is created at closed candle t:
# watch = {
#   "start_idx": idx_of_trigger_candle,
#   "expire_idx": start_idx + monitor_candles,
#   "long_level": float or None,
#   "short_level": float or None,
#   "atr": atr_value_at_trigger,
#   "trigger_time": pd.Timestamp of trigger candle
# }
# watches are checked on each new closed candle (by high/low)

# ---------------- Live Bot (paper trading) ----------------
class LiveDualEntryBot:
    def __init__(self, cfg: BotConfig):
        self.cfg = cfg
        self.balance = cfg.initial_balance
        self.watches: List[Dict] = []
        self.trades = []
        self.candles = pd.DataFrame(columns=['open','high','low','close','volume'])  # indexed by timestamp
        init_excel(self.cfg.logfile)

    async def start(self):
        client = await AsyncClient.create()
        bm = BinanceSocketManager(client)
        kline_socket = bm.kline_socket(self.cfg.pair, interval=self.cfg.interval)

        print(f"[INFO] Live forward-test dual-entry for {self.cfg.pair} ({self.cfg.interval})")
        async with kline_socket as stream:
            while True:
                res = await stream.recv()
                # print(res)
                # res is dict containing 'k' for kline
                k = res.get('k', {})
                is_closed = k.get('x', False)
                ts = pd.to_datetime(k.get('t'), unit='ms', utc=True)
                o = float(k.get('o')); h = float(k.get('h')); l = float(k.get('l')); c = float(k.get('c')); v = float(k.get('v'))
                # append/update last candle
                self._append_candle(ts, o, h, l, c, v)
                # only act when candle closed
                if is_closed:
                    # compute indicators
                    atr_series = compute_atr_from_df(self.candles, self.cfg.atr_period)
                    # ensure we have ATR value
                    if len(atr_series) >= self.cfg.atr_period:
                        current_atr = atr_series.iat[-1]
                    else:
                        current_atr = np.nan

                    # decide whether to create a watch (trigger) on this closed candle
                    if not np.isnan(current_atr) and current_atr >= self.cfg.min_atr:
                        last_close = self.candles['close'].iat[-1]
                        long_level = last_close + current_atr * self.cfg.level_mult
                        short_level = last_close - current_atr * self.cfg.level_mult
                        watch = {
                            "start_idx": len(self.candles)-1,
                            "expire_idx": len(self.candles)-1 + self.cfg.monitor_candles,
                            "long_level": long_level,
                            "short_level": short_level,
                            "atr": current_atr,
                            "trigger_time": self.candles.index[-1]
                        }
                        self.watches.append(watch)
                        print(f"[WATCH CREATED] {watch['trigger_time']} ATR={current_atr:.6f} long={long_level:.6f} short={short_level:.6f}")

                    # check existing watches with this new closed candle (the one that just closed)
                    self._process_watches_for_closed_candle()

    def _append_candle(self, ts, o, h, l, c, v):
        # add or replace last row with timestamp ts
        # keep buffer length limited
        self.candles.loc[ts] = [o, h, l, c, v]
        if len(self.candles) > self.cfg.candles_buffer:
            self.candles = self.candles.iloc[-self.cfg.candles_buffer:]

    def _process_watches_for_closed_candle(self):
        """
        Called after a candle closes. Evaluate each watch:
         - if high >= long_level => LONG triggered (entry at long_level)
         - elif low <= short_level => SHORT triggered (entry at short_level)
         - determine exit (TP/SL) using atr_at_trigger and monitor candle's high/low logic (like backtest)
         - remove expired watches
        """
        if self.candles.empty:
            return

        latest_idx = len(self.candles) - 1
        candle_high = self.candles['high'].iat[-1]
        candle_low = self.candles['low'].iat[-1]

        # we want to iterate watches copy because we may remove items
        new_watches = []
        for w in self.watches:
            # if already expired skip
            if latest_idx > w['expire_idx']:
                # expired without trigger
                continue

            # Determine whether this closed candle is within watch window (start exclusive -> we check subsequent candles)
            if latest_idx <= w['start_idx']:
                # this is the trigger candle itself, skip
                new_watches.append(w)
                continue

            triggered = False
            # LONG triggers if candle_high >= long_level
            if candle_high >= w['long_level']:
                triggered = True
                side = 'LONG'
                entry_price = w['long_level']
                atr_now = w['atr']
                # compute TP/SL
                tp_price = entry_price + atr_now * self.cfg.tp_atr_mult
                sl_price = entry_price - atr_now * self.cfg.sl_atr_mult
                # decide exit price using candle extremes and TP/SL (same logic as backtest)
                # if candle low <= sl_price -> SL hit, else if candle high >= tp_price -> TP hit
                if candle_low <= sl_price:
                    exit_price = sl_price
                    exit_reason = 'SL'
                elif candle_high >= tp_price:
                    exit_price = tp_price
                    exit_reason = 'TP'
                else:
                    # if neither hit within this candle, we cannot close yet (but backtest closed immediately when triggered)
                    # in backtest we used intrabar checks at the same candle; here we assume either TP or SL hit during that candle
                    # If neither, treat as no-exit yet and keep watch for next candle — but this would differ from backtest. To match backtest:
                    # We'll consider if high >= tp_price then TP, elif low <= sl_price then SL, else treat as no-exit yet and keep watching.
                    exit_price = None
                    exit_reason = None

            # SHORT triggers if candle_low <= short_level
            elif candle_low <= w['short_level']:
                triggered = True
                side = 'SHORT'
                entry_price = w['short_level']
                atr_now = w['atr']
                tp_price = entry_price - atr_now * self.cfg.tp_atr_mult
                sl_price = entry_price + atr_now * self.cfg.sl_atr_mult
                if candle_high >= sl_price:
                    exit_price = sl_price
                    exit_reason = 'SL'
                elif candle_low <= tp_price:
                    exit_price = tp_price
                    exit_reason = 'TP'
                else:
                    exit_price = None
                    exit_reason = None
            else:
                # not triggered on this candle
                new_watches.append(w)
                continue

            # If triggered and we have an exit price determined during same candle:
            if triggered and exit_price is not None:
                pnl = self._compute_pnl(entry_price, exit_price, side)
                self.balance += pnl
                trade = {
                    "pair": self.cfg.pair,
                    "watch_start": w['trigger_time'],
                    "trigger_side": side,
                    "trigger_level": entry_price,
                    "atr_at_trigger": atr_now,
                    "entry_time": w['trigger_time'],
                    "entry_price": entry_price,
                    "exit_time": self.candles.index[-1],
                    "exit_price": exit_price,
                    "pnl": pnl,
                    "exit_reason": exit_reason,
                    "balance_after": self.balance
                }
                self.trades.append(trade)
                append_trade_excel(self.cfg.logfile, [
                    trade['pair'], trade['watch_start'], trade['trigger_side'], trade['trigger_level'],
                    trade['atr_at_trigger'], trade['entry_time'], trade['entry_price'],
                    trade['exit_time'], trade['exit_price'], trade['pnl'],
                    trade['exit_reason'], trade['balance_after']
                ])
                print(f"[TRADE] {trade['exit_time']} {side} entry={entry_price:.6f} exit={exit_price:.6f} pnl={pnl:.6f} balance={self.balance:.4f}")
                # do not keep this watch
            else:
                # triggered but exit not yet determined -> keep watch for next candles (to try find TP/SL)
                # but we must ensure expire_idx still in future
                if latest_idx < w['expire_idx']:
                    new_watches.append(w)
                # else expired on same candle without exit -> drop
        self.watches = new_watches

    def _compute_pnl(self, entry_price: float, exit_price: float, side: str):
        """
        same PnL calc logic as backtest:
        - return fraction = (exit - entry)/entry for long; (entry - exit)/entry for short
        - pnl amount = fraction * balance * leverage - fee_rate*balance
        Note: fee applied once as in backtest formula (fee_rate * balance)
        """
        if side == 'LONG':
            frac = (exit_price - entry_price) / entry_price
        else:
            frac = (entry_price - exit_price) / entry_price
        pnl = frac * self.balance * self.cfg.leverage - self.cfg.fee_rate * self.balance
        return pnl

# ---------------- Run ----------------
if __name__ == "__main__":
    # ensure logfile exists and header set
    init_excel(cfg.logfile)
    bot = LiveDualEntryBot(cfg)
    try:
        asyncio.get_event_loop().run_until_complete(bot.start())
    except KeyboardInterrupt:
        print("[STOP] Interrupted by user. Trades saved to", cfg.logfile)
