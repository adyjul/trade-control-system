"""
live_bot_limit_scalper_m1.py

M1 scalping bot using LIMIT (maker) orders on Binance USDT-M Futures.
- Entry: place LIMIT order (price = reference +/- entry_offset) aiming to be maker
- If entry filled, place TP + SL as LIMIT (reduceOnly) orders
- If entry not filled within `entry_timeout_sec`, cancel it (missed trade)
- Filters: ATR minimum, cooldown between trades, min_tp_abs
- Position sizing via `max_exposure_frac` (fraction of balance*leverage)
- Testnet support and `live_mode` toggle

NOTES:
- Binance Futures behavior varies across API versions. Test thoroughly on testnet before enabling live_mode.
- Futures may not support strict POST_ONLY flags; we emulate maker intention by placing limit slightly away from market and canceling if immediate fill occurs.

Dependencies: python-binance (async), pandas, openpyxl, python-dotenv

Usage: set BINANCE_API_KEY/BINANCE_API_SECRET in .env or assign cfg.api_key / cfg.api_secret
Run: python3 live_bot_limit_scalper_m1.py
"""

import asyncio
import os
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import List, Dict, Optional

import numpy as np
import pandas as pd
from binance import AsyncClient, BinanceSocketManager
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()

# --------- Config ---------
@dataclass
class BotConfig:
    pair: str = "AVAXUSDT"
    interval: str = "1m"
    initial_balance: float = 15.0
    leverage: int = 20
    fee_rate: float = 0.0004
    min_atr: float = 0.0005
    atr_period: int = 14
    level_mult: float = 0.2
    tp_atr_mult: float = 0.9
    sl_atr_mult: float = 0.8
    monitor_candles: int = 3
    candles_buffer: int = 500
    min_hold_sec: int = 30
    logfile: str = "trades_log_limit.xlsx"
    live_mode: bool = True        # place orders when True
    api_key: Optional[str] = os.getenv("BINANCE_API_KEY")
    api_secret: Optional[str] = os.getenv("BINANCE_API_SECRET")
    testnet: bool = False

    # maker strategy settings
    entry_offset: float = 0.0005          # fraction of price to adjust limit to favor being maker (e.g. 0.0005 => 0.05%)
    entry_timeout_sec: int = 60           # cancel entry if not filled within X seconds
    min_tp_abs: float = 0.03              # minimal absolute TP (USDT) to avoid fee bleed
    max_exposure_frac: float = 0.1        # fraction of balance*leverage used per trade
    min_time_between_trades_sec: int = 10 # throttle openings to reduce overtrading

cfg = BotConfig()

# --------- Excel helpers ---------
def init_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append([
            "pair", "watch_start", "trigger_side", "trigger_level", "atr_at_trigger",
            "entry_time", "entry_price", "exit_time", "exit_price", "pnl",
            "exit_reason", "balance_after", "executed_entry_price", "executed_exit_price", "qty", "entry_order_id", "tp_order_id", "sl_order_id"
        ])
        wb.save(path)


def append_trade_excel(path: str, row: List):
    wb = load_workbook(path)
    ws = wb["Trades"]
    clean_row = []
    for v in row:
        if isinstance(v, datetime):
            clean_row.append(v.replace(tzinfo=None))
        else:
            clean_row.append(v)
    ws.append(clean_row)
    wb.save(path)

# --------- Indicators ---------

def compute_atr_from_df(df: pd.DataFrame, period: int):
    high_low = df['high'] - df['low']
    high_close = (df['high'] - df['close'].shift(1)).abs()
    low_close = (df['low'] - df['close'].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    return atr

# --------- Bot Implementation ---------
class LimitScalpBot:
    def __init__(self, cfg: BotConfig):
        self.cfg = cfg
        self.balance = cfg.initial_balance
        self.candles = pd.DataFrame(columns=['open','high','low','close','volume'])
        self.watches: List[Dict] = []
        self._current_position: Optional[Dict] = None
        self.client: Optional[AsyncClient] = None
        self._pending_entry_orders: Dict[str, Dict] = {}  # orderId -> metadata
        self._last_trade_time: Optional[pd.Timestamp] = None
        init_excel(self.cfg.logfile)

    async def _init_client(self):
        if self.client:
            return self.client
        if self.cfg.testnet:
            self.client = await AsyncClient.create(self.cfg.api_key, self.cfg.api_secret, testnet=True)
        else:
            self.client = await AsyncClient.create(self.cfg.api_key, self.cfg.api_secret)
        try:
            await self.client.futures_change_leverage(symbol=self.cfg.pair, leverage=self.cfg.leverage)
        except Exception as e:
            print("[WARN] set leverage failed:", e)
        return self.client

    async def _format_price(self, price: float) -> float:
        client = await self._init_client()
        try:
            info = await client.futures_exchange_info()
            for s in info['symbols']:
                if s['symbol'] == self.cfg.pair:
                    for f in s['filters']:
                        if f['filterType'] == 'PRICE_FILTER':
                            tick_size = float(f['tickSize'])
                            precision = int(round(-np.log10(tick_size)))
                            return round(price, precision)
        except Exception as e:
            print("[WARN] format price failed:", e)
        return price

    async def _close_client(self):
        if self.client:
            await self.client.close_connection()
            self.client = None

    async def _format_quantity(self, qty: float) -> float:
        client = await self._init_client()
        try:
            info = await client.futures_exchange_info()
            for s in info['symbols']:
                if s['symbol'] == self.cfg.pair:
                    for f in s['filters']:
                        if f['filterType'] == 'LOT_SIZE':
                            step_size = float(f['stepSize'])
                            min_qty = float(f['minQty'])
                            precision = int(round(-np.log10(step_size)))
                            qty = max(min_qty, round(qty, precision))
                            return qty
        except Exception as e:
            print("[WARN] format qty failed:", e)
        return qty

    async def _place_limit_order(self, side: str, price: float, qty: float, reduce_only: bool=False):
        """Place a LIMIT order. Returns order dict or None in paper-mode."""
        if not self.cfg.live_mode:
            # emulate an order id for paper-mode
            return {
                "orderId": f"paper-{int(datetime.utcnow().timestamp()*1000)}",
                "status": "NEW",
                "price": str(price),
                "origQty": str(qty),
                "avgPrice": str(price)
            }

        client = await self._init_client()
        try:
            params = {
                "symbol": self.cfg.pair,    # <-- use cfg.pair (fix)
                "side": side,
                "type": "LIMIT",
                "timeInForce": "GTC",
                "quantity": qty,
                "price": str(price),
            }
            if reduce_only:
                params["reduceOnly"] = True

            # use futures_create_order for futures endpoints
            order = await client.futures_create_order(**params)
            return order

        except Exception as e:
            # helpful debug log
            print(f"[ERROR] place_limit_order failed: {e}")
            return None

    async def _cancel_order(self, order_id: str):
        """Safe cancel: skip paper ids and empty ids."""
        if not order_id:
            return False
        # if it's a paper-mode id, nothing to cancel
        if str(order_id).startswith("paper-"):
            return True
        if not self.cfg.live_mode:
            return True
        client = await self._init_client()
        try:
            await client.futures_cancel_order(symbol=self.cfg.pair, orderId=order_id)
            return True
        except Exception as e:
            print('[WARN] cancel order failed:', e)
            return False
    
    async def place_tp_order(self, side: str, price: float, qty: float):
        if qty <= 0:
            return None
        try:
            return await self._place_limit_order(side, price, qty, reduce_only=True)
        except Exception as e:
            print(f"[WARN] TP order failed, fallback paper: {e}")
            return {"orderId": f"paper-tp-{int(datetime.utcnow().timestamp()*1000)}"}

    async def place_sl_order(self, side: str, price: float, qty: float):
        if qty <= 0:
            return None
        try:
            return await self._place_limit_order(side, price, qty, reduce_only=True)
        except Exception as e:
            print(f"[WARN] SL order failed, fallback paper: {e}")
            return {"orderId": f"paper-sl-{int(datetime.utcnow().timestamp()*1000)}"}

    async def _get_order(self, order_id: str):
        if not self.cfg.live_mode:
            return None
        client = await self._init_client()
        try:
            resp = await client.futures_get_order(symbol=self.cfg.pair, orderId=order_id)
            return resp
        except Exception as e:
            print('[WARN] get_order failed:', e)
            return None

    def _append_candle(self, ts, o, h, l, c, v):
        self.candles.loc[ts] = [o, h, l, c, v]
        if len(self.candles) > self.cfg.candles_buffer:
            self.candles = self.candles.iloc[-self.cfg.candles_buffer:]

    async def start(self):
        # initialize client for socket even if paper-mode
        await self._init_client()
        bm = BinanceSocketManager(self.client)
        kline_socket = bm.kline_socket(self.cfg.pair, interval=self.cfg.interval)
        print(f"[INFO] starting limit scalper {self.cfg.pair} interval={self.cfg.interval} live_mode={self.cfg.live_mode}")
        async with kline_socket as stream:
            while True:
                res = await stream.recv()
                k = res.get('k', {})
                is_closed = k.get('x', False)
                ts = pd.to_datetime(k.get('t'), unit='ms', utc=True)
                o = float(k.get('o')); h = float(k.get('h')); l = float(k.get('l')); c = float(k.get('c')); v = float(k.get('v'))
                self._append_candle(ts, o, h, l, c, v)

                # periodically check pending entry orders for timeout
                await self._check_pending_entries()

                if is_closed:
                    atr_series = compute_atr_from_df(self.candles, self.cfg.atr_period)
                    current_atr = atr_series.iat[-1] if len(atr_series) >= self.cfg.atr_period else np.nan
                    if not np.isnan(current_atr) and current_atr >= self.cfg.min_atr:
                        if self._current_position is None:
                            last_close = self.candles['close'].iat[-1]
                            long_level = last_close + current_atr * self.cfg.level_mult
                            short_level = last_close - current_atr * self.cfg.level_mult
                            watch = {
                                'start_idx': len(self.candles)-1,
                                'expire_idx': len(self.candles)-1 + self.cfg.monitor_candles,
                                'long_level': long_level,
                                'short_level': short_level,
                                'atr': current_atr,
                                'trigger_time': self.candles.index[-1]
                            }
                            self.watches.append(watch)
                            print(f"[WATCH] {watch['trigger_time']} ATR={current_atr:.6f} long={long_level:.6f} short={short_level:.6f}")

                    await self._process_watches_and_place_limit_entries()
                    # await self._process_current_position_and_maybe_close()
                    # cek apakah ada entry yang sudah terisi → mulai tick mode
                    for oid, meta in list(self._pending_entry_orders.items()):
                        await self._wait_for_entry_fill(
                            order_id=oid,
                            side=meta['side'],
                            entry_price=meta['trigger_price'],
                            qty=meta['qty'],
                            tp_price=meta['tp_price'],
                            sl_price=meta['sl_price']
                        )

    async def _check_pending_entries(self):
        # remove / cancel pending entry orders that timed out
        if not self._pending_entry_orders:
            return
        now = datetime.now(timezone.utc)
        to_remove = []
        for oid, meta in list(self._pending_entry_orders.items()):
            created = meta['created_at']
            elapsed = (now - created).total_seconds()
            if elapsed >= self.cfg.entry_timeout_sec:
                print(f"[TIMEOUT] Cancelling entry order {oid} after {elapsed:.1f}s")
                await self._cancel_order(oid)
                to_remove.append(oid)
        for oid in to_remove:
            self._pending_entry_orders.pop(oid, None)

    async def _process_watches_and_place_limit_entries(self):
        # throttle openings
        if self._current_position is not None:
            return
        now_idx = len(self.candles)-1
        now_ts = self.candles.index[-1]
        if self._last_trade_time is not None:
            elapsed_since = (now_ts - self._last_trade_time).total_seconds()
            if elapsed_since < self.cfg.min_time_between_trades_sec:
                return

        candle_high = self.candles['high'].iat[-1]
        candle_low = self.candles['low'].iat[-1]

        new_watches = []
        for w in self.watches:
            if now_idx <= w['start_idx']:
                new_watches.append(w); continue

            triggered_side = None
            trigger_price = None
            atr_now = w['atr']

            if candle_high >= w['long_level']:
                triggered_side = 'LONG'; trigger_price = w['long_level']
            elif candle_low <= w['short_level']:
                triggered_side = 'SHORT'; trigger_price = w['short_level']

            if triggered_side:
                # compute TP/SL
                tp_price = trigger_price + atr_now * self.cfg.tp_atr_mult if triggered_side == 'LONG' else trigger_price - atr_now * self.cfg.tp_atr_mult
                sl_price = trigger_price - atr_now * self.cfg.sl_atr_mult if triggered_side == 'LONG' else trigger_price + atr_now * self.cfg.sl_atr_mult

                # ensure minimal TP absolute
                tp_abs = abs(tp_price - trigger_price)
                if tp_abs < self.cfg.min_tp_abs:
                    # bump tp outward
                    if triggered_side == 'LONG':
                        tp_price = trigger_price + self.cfg.min_tp_abs
                        sl_price = trigger_price - self.cfg.min_tp_abs
                    else:
                        tp_price = trigger_price - self.cfg.min_tp_abs
                        sl_price = trigger_price + self.cfg.min_tp_abs
                    tp_abs = abs(tp_price - trigger_price)

                if tp_abs < self.cfg.min_tp_abs:
                    print(f"[SKIP] computed TP {tp_abs:.6f} < min_tp_abs")
                    continue

                # compute quantity
                max_notional = self.balance * self.cfg.leverage * self.cfg.max_exposure_frac
                qty = max_notional / trigger_price if trigger_price > 0 else 0.000001
                qty = max(0.000001, qty)
                qty = await self._format_quantity(qty)

                # prepare limit entry price with offset to try be maker
                if triggered_side == 'LONG':
                    limit_price = trigger_price * (1 - self.cfg.entry_offset)
                    side = 'BUY'
                else:
                    limit_price = trigger_price * (1 + self.cfg.entry_offset)
                    side = 'SELL'

                # place entry limit order
                # resp = await self._place_limit_order(side, round(limit_price,8), qty, reduce_only=False)
                limit_price = await self._format_price(limit_price)
                resp = await self._place_limit_order(side, limit_price, qty, reduce_only=False)
                if resp is None:
                    print('[WARN] entry order failed to create')
                    continue

                order_id = str(resp.get('orderId') or resp.get('clientOrderId') or resp.get('orderId'))
                self._pending_entry_orders[order_id] = {
                    'side': triggered_side,
                    'trigger_price': trigger_price,
                    'limit_price': float(limit_price),
                    'qty': qty,
                    'tp_price': tp_price,
                    'sl_price': sl_price,
                    'created_at': datetime.now(timezone.utc)
                }
                print(f"[ENTRY_ORDER] id={order_id} side={triggered_side} limit={limit_price:.6f} qty={qty} tp={tp_price:.6f} sl={sl_price:.6f}")
                # do not re-add watch
            else:
                if now_idx < w['expire_idx']:
                    new_watches.append(w)
        self.watches = new_watches

        # after placing, check quickly if any pending entry got filled (paper-mode simulates fill immediately)
        await self._poll_pending_entries_and_handle_fills()

    # async def _poll_pending_entries_and_handle_fills(self):
    #     # iterate over pending entry orders and see if filled
    #     for oid, meta in list(self._pending_entry_orders.items()):
    #         # in live_mode, poll order status
    #         order_info = await self._get_order(oid) if self.cfg.live_mode else {'status':'FILLED', 'avgPrice': str(meta['limit_price'])}
    #         status = (order_info.get('status') if isinstance(order_info, dict) else None) or order_info
    #         if isinstance(status, str) and status.upper() == 'FILLED' or (isinstance(order_info, dict) and order_info.get('status') == 'FILLED'):
    #             # handle fill
    #             executed_price = None
    #             if isinstance(order_info, dict):
    #                 executed_price = float(order_info.get('avgPrice') or order_info.get('price') or meta['limit_price'])
    #             else:
    #                 executed_price = meta['limit_price']

    #             # set current position metadata
    #             self._current_position = {
    #                 'side': meta['side'],
    #                 'entry_price': meta['trigger_price'],
    #                 'executed_entry_price': executed_price,
    #                 'tp_price': meta['tp_price'],
    #                 'sl_price': meta['sl_price'],
    #                 'qty': meta['qty'],
    #                 'entry_time': datetime.now(timezone.utc),   # tz-aware
    #                 'entry_order_id': oid,
    #                 'tp_order_id': None,
    #                 'sl_order_id': None
    #             }
    #             # place TP and SL as limit reduceOnly orders
    #             # TP (take profit)
    #             tp_side = 'SELL' if meta['side']=='LONG' else 'BUY'
    #             sl_side = 'SELL' if meta['side']=='SHORT' else 'BUY'
    #             # tp_resp = await self._place_limit_order(tp_side, round(meta['tp_price'],8), meta['qty'], reduce_only=True)
    #             # sl_resp = await self._place_limit_order(sl_side, round(meta['sl_price'],8), meta['qty'], reduce_only=True)
                
    #             tp_price = await self._format_price(meta['tp_price'])
    #             sl_price = await self._format_price(meta['sl_price'])

    #             try:
    #                 tp_resp = await self._place_limit_order(tp_side, tp_price, meta['qty'], reduce_only=True)
    #             except Exception as e:
    #                 print(f"[WARN] TP order failed (ReduceOnly?): {e}")
    #                 tp_resp = {'orderId': f"paper-tp-{int(datetime.utcnow().timestamp()*1000)}"}
                
    #             try:
    #                 sl_resp = await self._place_limit_order(sl_side, sl_price, meta['qty'], reduce_only=True)
    #             except Exception as e:
    #                 print(f"[WARN] SL order failed (ReduceOnly?): {e}")
    #                 sl_resp = {'orderId': f"paper-sl-{int(datetime.utcnow().timestamp()*1000)}"}

    #             # tp_resp = await self._place_limit_order(tp_side, tp_price, meta['qty'], reduce_only=True)
    #             # sl_resp = await self._place_limit_order(sl_side, sl_price, meta['qty'], reduce_only=True)


    #             tp_oid = str(tp_resp.get('orderId') if isinstance(tp_resp, dict) else f"paper-tp-{int(datetime.utcnow().timestamp()*1000)}")
    #             sl_oid = str(sl_resp.get('orderId') if isinstance(sl_resp, dict) else f"paper-sl-{int(datetime.utcnow().timestamp()*1000)}")

    #             self._current_position['tp_order_id'] = tp_oid
    #             self._current_position['sl_order_id'] = sl_oid

    #             print(f"[FILLED] entry executed={executed_price:.6f} qty={meta['qty']} tp_oid={tp_oid} sl_oid={sl_oid}")

    #             # remove pending entry record
    #             self._pending_entry_orders.pop(oid, None)
    #             self._last_trade_time = datetime.utcnow()

    async def _poll_pending_entries_and_handle_fills(self):
        for oid, meta in list(self._pending_entry_orders.items()):
            # cek status order
            order_info = await self._get_order(oid) if self.cfg.live_mode else {'status':'FILLED', 'avgPrice': str(meta['limit_price'])}
            status = (order_info.get('status') if isinstance(order_info, dict) else None) or order_info
            if isinstance(status, str) and status.upper() == 'FILLED' or (isinstance(order_info, dict) and order_info.get('status') == 'FILLED'):
                executed_price = float(order_info.get('avgPrice') or order_info.get('price') or meta['limit_price']) if isinstance(order_info, dict) else meta['limit_price']

                self._current_position = {
                    'side': meta['side'],
                    'entry_price': meta['trigger_price'],
                    'executed_entry_price': executed_price,
                    'tp_price': meta['tp_price'],
                    'sl_price': meta['sl_price'],
                    'qty': meta['qty'],
                    # 'entry_time': datetime.utcnow(),
                    'entry_time': datetime.now(timezone.utc),
                    'entry_order_id': oid,
                    'tp_order_id': None,
                    'sl_order_id': None
                }

                tp_side = 'SELL' if meta['side']=='LONG' else 'BUY'
                sl_side = 'SELL' if meta['side']=='SHORT' else 'BUY'

                # ===== PRE-CHECK BEFORE TP/SL =====
                tp_order_id = None
                sl_order_id = None
                if self._current_position['qty'] > 0:
                    # TP
                    try:
                        tp_price = await self._format_price(meta['tp_price'])
                        tp_resp = await self.place_tp_order(tp_side, tp_price, self._current_position['qty'])
                        if tp_resp:
                            tp_order_id = str(tp_resp.get('orderId') or tp_resp.get('clientOrderId') or f"paper-tp-{int(datetime.utcnow().timestamp()*1000)}")
                        else:
                            tp_order_id = f"paper-tp-{int(datetime.utcnow().timestamp()*1000)}"
                    except Exception as e:
                        print(f"[WARN] TP order failed (ReduceOnly?): {e}")
                        tp_order_id = f"paper-tp-{int(datetime.utcnow().timestamp()*1000)}"

                    # SL
                    try:
                        sl_price = await self._format_price(meta['sl_price'])
                        sl_resp = await self.place_sl_order(sl_side, sl_price, self._current_position['qty'])
                        if sl_resp:
                            sl_order_id = str(sl_resp.get('orderId') or sl_resp.get('clientOrderId') or f"paper-sl-{int(datetime.utcnow().timestamp()*1000)}")
                        else:
                            sl_order_id = f"paper-sl-{int(datetime.utcnow().timestamp()*1000)}"
                    except Exception as e:
                        print(f"[WARN] SL order failed (ReduceOnly?): {e}")
                        sl_order_id = f"paper-sl-{int(datetime.utcnow().timestamp()*1000)}"

                self._current_position['tp_order_id'] = tp_order_id
                self._current_position['sl_order_id'] = sl_order_id

                print(f"[FILLED] entry executed={executed_price:.6f} qty={self._current_position['qty']} tp_oid={tp_order_id} sl_oid={sl_order_id}")

                self._pending_entry_orders.pop(oid, None)
                self._last_trade_time = datetime.now(timezone.utc)

    async def _process_current_position_and_maybe_close(self):
        """Detect TP/SL hit (based on candle high/low) and actually close on exchange.
           Only write [CLOSED] log when exchange confirmed the close (close_position returned executed price).
        """
        if self._current_position is None:
            return
        pos = self._current_position
        latest_candle = self.candles.iloc[-1]

        now = self.candles.index[-1]
        elapsed = (now - pos['entry_time']).total_seconds()

        exit_reason = None
        exit_price = None
        executed_exit_price = None

        # NOTE: use consistent keys: tp_price / sl_price
        side = pos.get("side")
        qty = pos.get("qty", 0)
        tp_price = pos.get("tp_price")
        sl_price = pos.get("sl_price")

        if elapsed >= self.cfg.min_hold_sec:
            if side == 'LONG':
                if tp_price is not None and latest_candle['high'] >= tp_price:
                    exit_price = tp_price; exit_reason = 'TP'
                elif sl_price is not None and latest_candle['low'] <= sl_price:
                    exit_price = sl_price; exit_reason = 'SL'
            else:  # SHORT
                if tp_price is not None and latest_candle['low'] <= tp_price:
                    exit_price = tp_price; exit_reason = 'TP'
                elif sl_price is not None and latest_candle['high'] >= sl_price:
                    exit_price = sl_price; exit_reason = 'SL'

        if exit_price is not None:
            # Try to close on exchange. prefer_limit=False to ensure it closes (use market fallback)
            executed = await self.close_position(side, qty, exit_price=exit_price, prefer_limit=False)

            if executed is None:
                # close failed — do not mark closed in log; keep position and warn
                print(f"[ERROR] close failed on exchange for {side} qty={qty} exit_price={exit_price}. Keeping position open.")
                return

            executed_exit_price = float(executed)

            used_entry_price = pos.get('executed_entry_price') or pos.get('entry_price')
            pnl = self._compute_pnl(used_entry_price, executed_exit_price, pos['side'], pos['qty'])
            self.balance += pnl

            trade = [
                self.cfg.pair, pos['entry_time'], pos['side'], pos['entry_price'], None,
                pos['entry_time'], used_entry_price, now, executed_exit_price, pnl,
                exit_reason, self.balance, pos.get('executed_entry_price'), executed_exit_price, pos['qty'], pos.get('entry_order_id'), pos.get('tp_order_id'), pos.get('sl_order_id')
            ]
            append_trade_excel(self.cfg.logfile, trade)
            print(f"[CLOSED] {now} {pos['side']} entry={used_entry_price:.6f} exit={executed_exit_price:.6f} pnl={pnl:.6f} balance={self.balance:.4f}")

            # clear current position only after confirmed close
            self._current_position = None
    
    async def _execute_close(self, side: str, qty: float, exit_price: float, exit_reason: str):
        """Wrapper untuk menutup posisi saat TP/SL kena dari tick mode."""
        executed_price = await self.close_position(side, qty, exit_price=exit_price, prefer_limit=False)
        if executed_price is None:
            print(f"[ERROR] Failed to close {side} position qty={qty} at {exit_price}")
            return

        # update balance dan log trade
        pos = self._current_position
        if pos is None:
            return

        used_entry_price = pos.get('executed_entry_price') or pos.get('entry_price')
        pnl = self._compute_pnl(used_entry_price, executed_price, side, qty)
        self.balance += pnl

        now = datetime.now(timezone.utc)
        trade = [
            self.cfg.pair, pos['entry_time'], side, pos['entry_price'], None,
            pos['entry_time'], used_entry_price, now, executed_price, pnl,
            exit_reason, self.balance, pos.get('executed_entry_price'), executed_price, qty,
            pos.get('entry_order_id'), pos.get('tp_order_id'), pos.get('sl_order_id')
        ]
        append_trade_excel(self.cfg.logfile, trade)
        print(f"[CLOSED] {now} {side} entry={used_entry_price:.6f} exit={executed_price:.6f} pnl={pnl:.6f} balance={self.balance:.4f}")

        # hapus posisi aktif
        self._current_position = None
    
    async def _wait_for_entry_fill(self, order_id, side, entry_price, qty, tp_price, sl_price):
        """Cek status order sampai FILLED, baru aktifkan socket close."""
        print(f"[WAIT] Menunggu order {order_id} terisi...")

        while True:
            try:
                order = await self.client.get_order(symbol=self.cfg.pair, orderId=order_id)
                status = order["status"]

                if status == "FILLED":
                    executed_entry_price = float(order["avgPrice"]) if order.get("avgPrice") else entry_price
                    entry_time = pd.Timestamp.utcnow()

                    # Simpan posisi aktif
                    self._current_position = {
                        "side": side,
                        "qty": qty,
                        "entry_price": entry_price,
                        "executed_entry_price": executed_entry_price,
                        "entry_time": entry_time,
                        "tp_price": tp_price,
                        "sl_price": sl_price,
                        "entry_order_id": order_id,
                    }

                    print(f"[FILLED] {side} qty={qty} @ {executed_entry_price}")
                    
                    # Mulai socket untuk pantau TP/SL real-time
                    await self.start_socket_for_close()
                    return True

                elif status in ("CANCELED", "REJECTED", "EXPIRED"):
                    print(f"[FAILED] Entry order {order_id} status={status}")
                    return False

                else:
                    # status masih NEW / PARTIALLY_FILLED → tunggu
                    await asyncio.sleep(1)

            except Exception as e:
                print(f"[ERROR] cek status order: {e}")
                await asyncio.sleep(2)
    
    async def start_socket_for_close(self):
       bm = BinanceSocketManager(self.client)

       async with bm.mark_price_socket(symbol=self.cfg.pair) as stream:
            print("[SOCKET] Listening mark price for close...")
            while self._current_position is not None:
                msg = await stream.recv()
                data = msg['data']
                price = float(data['p'])   # mark price

                # kirim price langsung, bukan msg dict
                self._on_price_tick(price)

    def _on_price_tick(self, price: float):
        try:
            if self._current_position is None:
                return

            pos = self._current_position
            side = pos["side"]
            qty = pos["qty"]
            tp_price = pos.get("tp_price")
            sl_price = pos.get("sl_price")

            exit_price = None
            exit_reason = None

            if side == "LONG":
                if tp_price and price >= tp_price:
                    exit_price = tp_price; exit_reason = "TP"
                elif sl_price and price <= sl_price:
                    exit_price = sl_price; exit_reason = "SL"
            else:  # SHORT
                if tp_price and price <= tp_price:
                    exit_price = tp_price; exit_reason = "TP"
                elif sl_price and price >= sl_price:
                    exit_price = sl_price; exit_reason = "SL"

            if exit_price is not None:
                asyncio.create_task(self._execute_close(side, qty, exit_price, exit_reason))

        except Exception as e:
            print(f"[ERROR] on_price_tick: {e}")

    async def close_position(self, side: str, qty: float, exit_price: Optional[float] = None, prefer_limit: bool = False):
        """
        Close a position on the exchange.
        - side: "LONG" or "SHORT" (bot internal)
        - qty: quantity to close
        - exit_price: prefered limit price (if prefer_limit True)
        - prefer_limit: try limit reduceOnly close first; if fails, use market
        Returns executed_price (float) on success, None on failure.
        """
        if qty <= 0:
            print("[WARN] close_position called with qty <= 0")
            return None

        close_side = 'SELL' if side == 'LONG' else 'BUY'

        # cancel existing TP/SL orders to avoid conflicts (safe: _cancel_order skips paper ids)
        try:
            if self._current_position:
                if self._current_position.get('tp_order_id'):
                    await self._cancel_order(self._current_position.get('tp_order_id'))
                if self._current_position.get('sl_order_id'):
                    await self._cancel_order(self._current_position.get('sl_order_id'))
        except Exception as e:
            print("[WARN] failed canceling tp/sl before close:", e)

        # PAPER MODE: emulate immediate close using last known price
        if not self.cfg.live_mode:
            executed = exit_price if exit_price is not None else float(self.candles['close'].iat[-1])
            return float(executed)

        client = await self._init_client()

        # helper to extract executed price from resp
        def _extract_price(resp):
            if not resp:
                return None
            # common fields
            for k in ('avgPrice', 'avgFilledPrice', 'price'):
                if isinstance(resp.get(k), (str, float)):
                    try:
                        return float(resp.get(k))
                    except:
                        pass
            # tries fills if available
            fills = resp.get('fills') or resp.get('fill') or []
            if fills and isinstance(fills, list):
                num = 0.0; den = 0.0
                for f in fills:
                    p = float(f.get('price', 0))
                    q = float(f.get('qty', f.get('qty', 0)))
                    num += p * q; den += q
                if den > 0:
                    return num/den
            return None

        # Try preferred route: LIMIT reduceOnly at exit_price (to remain maker) if requested
        if prefer_limit and exit_price is not None:
            try:
                px = await self._format_price(exit_price)
                resp = await client.futures_create_order(
                    symbol=self.cfg.pair,
                    side=close_side,
                    type='LIMIT',
                    timeInForce='GTC',
                    quantity=qty,
                    price=str(px),
                    reduceOnly=True
                )
                exec_px = _extract_price(resp)
                # if order created but not filled immediately, we should wait/poll — but for safety, fallback to market if not filled
                if exec_px is not None:
                    # quick verification
                    await asyncio.sleep(0.2)
                    return exec_px
                else:
                    # place market fallback
                    print("[INFO] limit-close created but not filled immediately -> fallback to MARKET close")
            except Exception as e:
                print("[WARN] limit reduceOnly close failed:", e)

        # MARKET close (guaranteed fill). Use reduceOnly if exchange supports for MARKET; otherwise market without reduceOnly.
        try:
            # try with reduceOnly param first (some API versions support it)
            try:
                resp = await client.futures_create_order(
                    symbol=self.cfg.pair,
                    side=close_side,
                    type='MARKET',
                    quantity=qty,
                    reduceOnly=True
                )
            except Exception:
                # fallback: market without reduceOnly (still should close)
                resp = await client.futures_create_order(
                    symbol=self.cfg.pair,
                    side=close_side,
                    type='MARKET',
                    quantity=qty
                )

            exec_px = _extract_price(resp)
            # small pause and verify position actually zeroed
            await asyncio.sleep(0.3)
            try:
                pos_info = await client.futures_position_information(symbol=self.cfg.pair)
                # pos_info usually list of dicts
                pos_amt = None
                if isinstance(pos_info, list):
                    for p in pos_info:
                        if p.get('symbol') == self.cfg.pair:
                            pos_amt = float(p.get('positionAmt', 0))
                            break
                elif isinstance(pos_info, dict):
                    pos_amt = float(pos_info.get('positionAmt', 0))

                if pos_amt is not None and abs(pos_amt) > 1e-8:
                    print(f"[WARN] close_position: positionAmt still not zero after close ({pos_amt}).")
                return exec_px if exec_px is not None else float(self.candles['close'].iat[-1])
            except Exception as e:
                print("[WARN] unable to verify position after close:", e)
                return exec_px if exec_px is not None else float(self.candles['close'].iat[-1])

        except Exception as e:
            print("[ERROR] close_position failed:", e)
            return None

    def _compute_pnl(self, entry_price: float, exit_price: float, side: str, qty: float):
        if side == 'LONG':
            price_diff = exit_price - entry_price
        else:
            price_diff = entry_price - exit_price
        gross_pnl = price_diff * qty
        entry_notional = entry_price * qty
        exit_notional = exit_price * qty
        total_fee = self.cfg.fee_rate * (entry_notional + exit_notional)
        pnl = gross_pnl - total_fee
        return pnl

# --------- run ---------
if __name__ == '__main__':
    cfg.testnet = False
    cfg.live_mode = True
    cfg.api_key = os.getenv('BINANCE_API_KEY')
    cfg.api_secret = os.getenv('BINANCE_API_SECRET')

    init_excel(cfg.logfile)
    bot = LimitScalpBot(cfg)
    try:
        asyncio.get_event_loop().run_until_complete(bot.start())
    except KeyboardInterrupt:
        print('[STOP] interrupted. Trades saved to', cfg.logfile)
    finally:
        try:
            asyncio.get_event_loop().run_until_complete(bot._close_client())
        except:
            pass
